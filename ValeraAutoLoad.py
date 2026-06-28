import sys
import os
import io
import re
import gc
import time
import shutil
import logging
import traceback
import warnings
from logging.handlers import RotatingFileHandler
from dataclasses import dataclass, field
from typing import Optional, List, Dict, Tuple
from urllib.parse import urlparse, quote, urljoin
from collections import deque, defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from html import escape

import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

from PIL import Image as PILImage, ImageOps
import openpyxl
from openpyxl.styles import PatternFill

from PySide6.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QLineEdit, QCheckBox, QComboBox, QMessageBox, QGridLayout,
    QToolButton, QFileDialog, QGroupBox, QTableWidget,
    QTableWidgetItem, QAbstractItemView, QHeaderView, QDialog, QFrame,
    QScrollArea, QSizePolicy, QProgressBar
)
from PySide6.QtCore import QThread, Signal, QSettings, Qt, QRect, QTimer, QMutex, QWaitCondition, QStandardPaths, QUrl
from PySide6.QtGui import (
    QPixmap, QPainter, QColor, QBrush, QPen, QPalette, QDesktopServices,
    QKeySequence, QShortcut, QIcon, QLinearGradient, QIntValidator
)

warnings.filterwarnings("ignore", message="Unverified HTTPS request")

try:
    import numpy as np
    HAS_NUMPY = True
except ImportError:
    HAS_NUMPY = False

# ============================================================
# LOGGING
# ============================================================
LOGGER = logging.getLogger("Valera")
LOGGER.setLevel(logging.DEBUG)


def setup_logging():
    try:
        app_dir = QStandardPaths.writableLocation(QStandardPaths.AppDataLocation)
    except Exception:
        app_dir = os.path.expanduser("~")
    log_dir = os.path.join(app_dir, "ValeraSoft", "logs")
    os.makedirs(log_dir, exist_ok=True)
    handler = RotatingFileHandler(
        os.path.join(log_dir, "valera.log"),
        maxBytes=5 * 1024 * 1024, backupCount=3, encoding="utf-8"
    )
    handler.setFormatter(logging.Formatter(
        "%(asctime)s [%(levelname)s] %(message)s", datefmt="%Y-%m-%d %H:%M:%S"
    ))
    if not LOGGER.handlers:
        LOGGER.addHandler(handler)


def apply_light_theme(app):
    app.setStyle("Fusion")
    p = QPalette()
    p.setColor(QPalette.Window, QColor(240, 240, 240))
    p.setColor(QPalette.WindowText, Qt.black)
    p.setColor(QPalette.Base, QColor(255, 255, 255))
    p.setColor(QPalette.AlternateBase, QColor(245, 245, 245))
    p.setColor(QPalette.ToolTipBase, QColor(255, 255, 220))
    p.setColor(QPalette.ToolTipText, Qt.black)
    p.setColor(QPalette.Text, Qt.black)
    p.setColor(QPalette.Button, QColor(240, 240, 240))
    p.setColor(QPalette.ButtonText, Qt.black)
    p.setColor(QPalette.BrightText, Qt.red)
    p.setColor(QPalette.Link, QColor(42, 130, 218))
    p.setColor(QPalette.Highlight, QColor(42, 130, 218))
    p.setColor(QPalette.HighlightedText, Qt.white)
    app.setPalette(p)


# ============================================================
# CONSTANTS & CONFIG
# ============================================================
class Defaults:
    THRESHOLD_WHITE = 255
    THRESHOLD_NEAR_DELTA = 15
    MIN_BG_RATIO = 0.12
    QUESTIONABLE_BG_RATIO = 0.03
    MAX_UNDO_STEPS = 15
    LOG_MAX_BYTES = 5 * 1024 * 1024
    LOG_BACKUP_COUNT = 3
    PREVIEW_PAGE_SIZE = 30
    DOWNLOAD_WORKERS = 2
    JPEG_QUALITY = 92
    WEBP_QUALITY = 90
    PDF_DPI = 200
    THUMBNAIL_SIZE = 50
    PROGRESS_VALERA_SIZE = 48
    MAX_PX_CAP = 16000
    MAX_UPSCALE_CAP = 1000
    MIN_PADDING = 2
    WM_OPACITY = 0.5
    WM_MAX_RATIO = 4
    WM_MIN_SIZE = 50


@dataclass
class ProcessingPreset:
    name: str
    min_px: int
    max_px: int
    max_upscale_pct: int
    align: str
    fmt: str
    center_square: bool
    padding_pct: int
    remove_meta: bool = False
    replace_transparent: bool = True
    process_white_bg: bool = True


PRESETS = {
    "Пользовательский": ProcessingPreset(
        "Пользовательский", 0, 4000, 50, "height", "jpg", True, 10, False, True, True
    ),
    "santehnica.ru": ProcessingPreset(
        "santehnica.ru", 1080, 4000, 100, "height", "jpg", True, 10, True, True, True
    ),
    "A4": ProcessingPreset(
        "A4", 2480, 3508, 20, "height", "png", False, 0, True, False, False
    ),
}


@dataclass
class AppSettings:
    source: str = ""
    source_type: str = "Excel"
    out_dir: str = ""
    article_col: str = "A"
    url_from: str = "B"
    url_to: str = "P"
    rename_mode: str = "article"
    folder_sort: str = "В порядке Excel"
    min_px: int = 0
    max_px: int = 4000
    max_upscale_pct: int = 50
    align: str = "height"
    fmt: str = "jpg"
    center_square: bool = True
    padding_pct: int = 10
    replace_transparent: bool = True
    remove_meta: bool = False
    report_copy: bool = False
    clean_raw: bool = False
    ssl_verify: bool = True
    aggressive_parse: bool = True
    preset_name: str = "santehnica.ru"
    pdf_always_square: bool = True
    selected_rejected_files: List[str] = field(default_factory=list)
    white_threshold: int = Defaults.THRESHOLD_WHITE
    watermark_path: str = ""
    process_white_bg: bool = True


# ============================================================
# HELPERS
# ============================================================
URL_REGEX = re.compile(r'(https?://[^\s,;]+|www\.[^\s,;]+)')

SKIP_EXTENSIONS = {
    '.xlsx', '.xls', '.csv', '.doc', '.docx', '.txt', '.rtf',
    '.ods', '.tmp', '.bak', '.ini', '.cfg', '.log', '.zip',
    '.rar', '.7z', '.exe', '.msi', '.dll', '.bat', '.cmd'
}
VALID_EXTENSIONS = ('.png', '.jpg', '.jpeg', '.webp', '.pdf')


def col_letter_to_index(letter: str) -> int:
    col = 0
    for ch in letter.upper():
        col = col * 26 + (ord(ch) - ord("A") + 1)
    return col - 1


def generate_columns(count: int = 120) -> List[str]:
    res = []
    for i in range(count):
        idx = i + 1
        chars = []
        while idx > 0:
            idx, rem = divmod(idx - 1, 26)
            chars.append(chr(rem + ord("A")))
        res.append("".join(reversed(chars)))
    return res


def is_excel_locked(fp: str) -> bool:
    if not os.path.exists(fp):
        return False
    try:
        with open(fp, "rb+"):
            pass
        return False
    except (OSError, PermissionError, IOError):
        return True


def is_processable_file(filename: str) -> bool:
    fn = filename.lower()
    if fn.startswith(("~$", ".")):
        return False
    if fn.endswith(tuple(SKIP_EXTENSIONS)):
        return False
    return fn.endswith(VALID_EXTENSIONS)


def extract_urls(val) -> List[str]:
    if not val:
        return []
    urls = URL_REGEX.findall(str(val).strip())
    cleaned = []
    for u in urls:
        u = u.rstrip('.,;:')
        if u.startswith('www.'):
            u = 'https://' + u
        if u.startswith('http'):
            cleaned.append(u)
    return cleaned


def sanitize_filename(name: str, max_len: int = 100) -> str:
    name = re.sub(r'[<>:"/\\|?*\x00-\x1f]', '_', name)
    return name.strip('. ')[:max_len] or "unnamed"


def detect_ext_from_bytes(data: bytes) -> Optional[str]:
    if len(data) >= 8 and data[:8] == b'\x89PNG\r\n\x1a\n':
        return ".png"
    if len(data) >= 12 and data[:4] == b'RIFF' and data[8:12] == b'WEBP':
        return ".webp"
    if len(data) >= 5 and data[:5] == b'%PDF-':
        return ".pdf"
    return None


def ensure_rgb(img, bg: tuple = (255, 255, 255)):
    if img.mode in ("RGBA", "P"):
        rgb = PILImage.new("RGB", img.size, bg)
        if img.mode == "RGBA":
            rgb.paste(img, mask=img.split()[3])
        else:
            rgb.paste(img)
        return rgb
    return img.convert("RGB")


# ============================================================
# UNDO MANAGER
# ============================================================
class UndoManager:
    def __init__(self, max_steps: int = Defaults.MAX_UNDO_STEPS):
        self.history = deque(maxlen=max_steps)
        self.temp_dir = None

    def init_temp(self, path: str):
        self.temp_dir = path
        os.makedirs(self.temp_dir, exist_ok=True)

    def backup_for_edit(self, filepath: str) -> Optional[str]:
        if not self.temp_dir:
            return None
        ts = time.strftime("%Y%m%d_%H%M%S") + f"_{os.urandom(2).hex()}"
        backup = os.path.join(self.temp_dir, f"{ts}_{os.path.basename(filepath)}")
        try:
            shutil.copy2(filepath, backup)
            return backup
        except (OSError, shutil.Error) as e:
            LOGGER.error("Undo backup edit failed: %s", e)
            return None

    def backup_for_delete(self, filepath: str) -> Optional[str]:
        if not self.temp_dir:
            return None
        ts = time.strftime("%Y%m%d_%H%M%S") + f"_{os.urandom(2).hex()}"
        backup = os.path.join(self.temp_dir, f"{ts}_{os.path.basename(filepath)}")
        try:
            os.makedirs(os.path.dirname(backup), exist_ok=True)
            shutil.move(filepath, backup)
            return backup
        except (OSError, shutil.Error) as e:
            LOGGER.error("Undo backup delete failed: %s", e)
            return None

    def push_batch(self, operations: List[Tuple[str, str, str]]):
        if operations:
            self.history.append(operations)

    def undo(self) -> Tuple[bool, str]:
        if not self.history:
            return False, "Нет действий для отмены"
        batch = self.history.pop()
        try:
            for op in reversed(batch):
                action, orig, backup = op
                if action == "edit":
                    shutil.copy2(backup, orig)
                elif action == "delete":
                    os.makedirs(os.path.dirname(orig), exist_ok=True)
                    shutil.move(backup, orig)
            return True, f"Отменено {len(batch)} действий"
        except (OSError, shutil.Error) as e:
            return False, f"Ошибка при отмене: {e}"

    def clear(self):
        self.history.clear()
        if self.temp_dir and os.path.exists(self.temp_dir):
            try:
                shutil.rmtree(self.temp_dir)
            except OSError:
                pass


# ============================================================
# IMAGE PROCESSOR
# ============================================================
class ImageProcessor:
    def classify_white_background(self, img: PILImage.Image, threshold: int = Defaults.THRESHOLD_WHITE) -> str:
        img_rgba = img.convert("RGBA") if img.mode != "RGBA" else img.copy()
        img_small = img_rgba.copy()
        img_small.thumbnail((100, 100))
        w, h = img_small.size
        if w < 5 or h < 5:
            return "interior"

        px = img_small.load()
        is_pure = [[False] * w for _ in range(h)]
        is_near = [[False] * w for _ in range(h)]
        near_thr = max(threshold - Defaults.THRESHOLD_NEAR_DELTA, 0)

        for y in range(h):
            for x in range(w):
                r, g, b, a = px[x, y]
                if a == 0 or (a == 255 and r >= threshold and g >= threshold and b >= threshold):
                    is_pure[y][x] = True
                elif a == 255 and r >= near_thr and g >= near_thr and b >= near_thr:
                    is_near[y][x] = True

        border = set()
        for x in range(w):
            border.update([(x, 0), (x, h - 1)])
        for y in range(1, h - 1):
            border.update([(0, y), (w - 1, y)])

        vis = [[False] * w for _ in range(h)]
        pure = 0
        for bx, by in border:
            if is_pure[by][bx] and not vis[by][bx]:
                q = deque([(bx, by)])
                vis[by][bx] = True
                while q:
                    cx, cy = q.popleft()
                    pure += 1
                    for dx, dy in [(-1, 0), (1, 0), (0, -1), (0, 1)]:
                        nx, ny = cx + dx, cy + dy
                        if 0 <= nx < w and 0 <= ny < h and not vis[ny][nx] and is_pure[ny][nx]:
                            vis[ny][nx] = True
                            q.append((nx, ny))

        total = w * h
        if total > 0 and (pure / total) >= Defaults.MIN_BG_RATIO:
            return "white_bg"

        vis2 = [[False] * w for _ in range(h)]
        near = 0
        for bx, by in border:
            if (is_pure[by][bx] or is_near[by][bx]) and not vis2[by][bx]:
                q = deque([(bx, by)])
                vis2[by][bx] = True
                while q:
                    cx, cy = q.popleft()
                    near += 1
                    for dx, dy in [(-1, 0), (1, 0), (0, -1), (0, 1)]:
                        nx, ny = cx + dx, cy + dy
                        if 0 <= nx < w and 0 <= ny < h and not vis2[ny][nx] and (is_pure[ny][nx] or is_near[ny][nx]):
                            vis2[ny][nx] = True
                            q.append((nx, ny))

        if total > 0 and (near / total) >= Defaults.MIN_BG_RATIO:
            return "questionable"
        if total > 0 and Defaults.QUESTIONABLE_BG_RATIO <= (pure / total) < Defaults.MIN_BG_RATIO:
            return "questionable"
        return "interior"

    def _get_content_bbox(self, img: PILImage.Image, threshold: int = Defaults.THRESHOLD_WHITE):
        img_rgba = img.convert("RGBA") if img.mode != "RGBA" else img.copy()
        w, h = img_rgba.size
        if w == 0 or h == 0:
            return None

        if HAS_NUMPY:
            try:
                arr = np.array(img_rgba)
                mask = (arr[:, :, 3] > 0) & (
                    (arr[:, :, 0] < threshold) |
                    (arr[:, :, 1] < threshold) |
                    (arr[:, :, 2] < threshold)
                )
                rows = np.any(mask, axis=1)
                cols = np.any(mask, axis=0)
                if not rows.any() or not cols.any():
                    return None
                left = int(cols.argmax())
                top = int(rows.argmax())
                right = int(len(cols) - cols[::-1].argmax())
                bottom = int(len(rows) - rows[::-1].argmax())
                if right <= left or bottom <= top:
                    return None
                return (left, top, right, bottom)
            except Exception:
                pass

        data = list(img_rgba.getdata())
        mask_data = bytearray(w * h)
        for i, (r, g, b, a) in enumerate(data):
            if a == 0 or (a == 255 and r >= threshold and g >= threshold and b >= threshold):
                mask_data[i] = 0
            else:
                mask_data[i] = 255
        mask = PILImage.frombytes("L", (w, h), bytes(mask_data))
        return mask.getbbox()

    def _crop_to_content(self, img: PILImage.Image, threshold: int = Defaults.THRESHOLD_WHITE):
        bbox = self._get_content_bbox(img, threshold)
        if bbox and (bbox[2] - bbox[0]) > 0 and (bbox[3] - bbox[1]) > 0:
            return img.crop(bbox)
        return None

    def _center_in_square(self, img: PILImage.Image, padding_pct: int):
        w, h = img.size
        if w == 0 or h == 0:
            return img
        max_side = max(w, h)
        pad = int(max_side * (padding_pct / 100.0))
        cs = max_side + 2 * pad
        bg = PILImage.new("RGBA", (cs, cs), (255, 255, 255, 255))
        if img.mode == "RGBA":
            bg.paste(img, ((cs - w) // 2, (cs - h) // 2), img)
        else:
            bg.paste(img, ((cs - w) // 2, (cs - h) // 2))
        return bg

    def _smart_crop_to_square(self, img: PILImage.Image, threshold: int = Defaults.THRESHOLD_WHITE):
        w_orig, h_orig = img.size
        bbox = self._get_content_bbox(img, threshold)
        if not bbox:
            return self._center_in_square(img, 0)
        left, upper, right, lower = bbox
        img_c = img.crop(bbox)
        w_c, h_c = img_c.size
        s = max(w_c, h_c)
        canvas = PILImage.new("RGBA", (s, s), (255, 255, 255, 255))
        x = (s - w_c) // 2
        y = (s - h_c) // 2
        touched_top = (upper == 0)
        touched_bottom = (lower == h_orig)
        touched_left = (left == 0)
        touched_right = (right == w_orig)
        if touched_top:
            y = 0
        if touched_bottom:
            y = s - h_c
        if touched_left:
            x = 0
        if touched_right:
            x = s - w_c
        if touched_top and touched_bottom:
            y = (s - h_c) // 2
        if touched_left and touched_right:
            x = (s - w_c) // 2
        if img_c.mode == "RGBA":
            canvas.paste(img_c, (x, y), img_c)
        else:
            canvas.paste(img_c, (x, y))
        return canvas

    def apply_watermark(self, img: PILImage.Image, watermark_path: str):
        if not watermark_path or not os.path.exists(watermark_path):
            return img
        try:
            wm = PILImage.open(watermark_path).convert("RGBA")
            iw, ih = img.size
            wm_max_w = max(iw // Defaults.WM_MAX_RATIO, Defaults.WM_MIN_SIZE)
            wm_max_h = max(ih // Defaults.WM_MAX_RATIO, Defaults.WM_MIN_SIZE)
            wm.thumbnail((wm_max_w, wm_max_h), PILImage.LANCZOS)
            wm_w, wm_h = wm.size
            padding = max(iw // 50, 10)
            pos_x = iw - wm_w - padding
            pos_y = ih - wm_h - padding
            alpha = wm.split()[3]
            alpha = alpha.point(lambda p: int(p * Defaults.WM_OPACITY))
            wm.putalpha(alpha)
            if img.mode != "RGBA":
                img = img.convert("RGBA")
            img.paste(wm, (pos_x, pos_y), wm)
            wm.close()
            return img
        except Exception as e:
            LOGGER.error("Watermark error: %s", e)
            return img

    def process(self, img: PILImage.Image, s: AppSettings, bg_type: str = "white_bg",
                skip_center: bool = False, threshold: int = Defaults.THRESHOLD_WHITE):
        if s.remove_meta:
            img = ImageOps.exif_transpose(img)
            img.info.pop("exif", None)
        
        # Если обработка белого фона отключена — пропускаем центрирование и обрезку
        if s.process_white_bg and s.center_square and bg_type == "white_bg" and not skip_center:
            cropped = self._crop_to_content(img, threshold)
            if cropped is not None:
                img = cropped
            if img.size[0] > 0 and img.size[1] > 0:
                img = self._center_in_square(img, s.padding_pct)
        
        w, h = img.size
        if w == 0 or h == 0:
            return PILImage.new("RGBA", (1, 1), (255, 255, 255, 255)), "!ОШИБКА_"
        if s.replace_transparent and img.mode == "RGBA":
            bg = PILImage.new("RGB", img.size, (255, 255, 255))
            bg.paste(img, mask=img.split()[3])
            img = bg
        td = h if s.align == "height" else w
        if td == 0:
            td = 1
        scale = 1.0
        pref = ""
        if s.min_px > 0 and td < s.min_px:
            pct = ((s.min_px - td) / td) * 100
            scale = s.min_px / td
            if s.max_upscale_pct > 0 and pct > s.max_upscale_pct:
                pref = "!РАЗМЕР_"
        elif s.max_px > 0 and td > s.max_px:
            scale = s.max_px / td
        if scale != 1.0:
            img = img.resize((int(w * scale), int(h * scale)), PILImage.LANCZOS)
        if s.watermark_path:
            img = self.apply_watermark(img, s.watermark_path)
        return img, pref

    def process_with_analysis(self, img: PILImage.Image, s: AppSettings, force_type: str = "",
                              threshold: int = Defaults.THRESHOLD_WHITE):
        if force_type == "WHITE_BG":
            cropped = self._crop_to_content(img, threshold)
            if cropped is not None:
                img = cropped
            if img.size[0] > 0 and img.size[1] > 0 and s.center_square and s.process_white_bg:
                img = self._center_in_square(img, s.padding_pct)
            p, _ = self.process(img, s, "white_bg", threshold=threshold)
            return p, "", "FORCED_WHITE"
        if force_type == "PDF_SQUARE":
            cropped = self._crop_to_content(img, threshold)
            if cropped is not None:
                img = cropped
            if img.size[0] > 0 and img.size[1] > 0 and s.center_square and s.process_white_bg:
                img = self._center_in_square(img, s.padding_pct)
            p, _ = self.process(img, s, "white_bg", threshold=threshold)
            return p, "", "OK_PDF"
        if force_type == "SIZE":
            p, _ = self.process(img, s, "white_bg", threshold=threshold)
            return p, "", "FORCED_SIZE"
        
        # Если обработка белого фона отключена — считаем всё interior
        if not s.process_white_bg:
            p, sd = self.process(img, s, "interior", threshold=threshold)
            return p, sd, "OK_INTERIOR"
        
        bg_type = self.classify_white_background(img, threshold)
        p, sd = self.process(img, s, bg_type, threshold=threshold)
        if sd:
            return p, sd, "SIZE_FAIL"
        if bg_type == "questionable":
            return p, "!БЕЛЫЙ_", "QUESTION_WHITE"
        if bg_type == "white_bg":
            return p, "", "OK_WHITE"
        return p, "", "OK_INTERIOR"


# ============================================================
# EXCEL MANAGER
# ============================================================
class ExcelManager:
    def __init__(self, filepath: str):
        self.filepath = filepath
        self.wb = None

    def load(self):
        self.wb = openpyxl.load_workbook(self.filepath)
        return self.wb

    def close(self):
        if self.wb:
            self.wb.close()
            self.wb = None

    def count_urls(self, url_cols, article_col: int) -> int:
        total = 0
        for sheet in self.wb.worksheets:
            max_col = max(article_col, max(url_cols)) + 1
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, max_col=max_col):
                if article_col >= len(row) or not row[article_col].value:
                    continue
                for c in url_cols:
                    if c < len(row) and row[c].value:
                        total += len(extract_urls(row[c].value))
        return total


# ============================================================
# CLOUD RESOLVER
# ============================================================
class CloudResolver:
    def __init__(self, verify_ssl: bool = True):
        self.session = requests.Session()
        self.session.headers.update({
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
            "Accept-Language": "ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7",
            "Origin": "https://cloud.mail.ru",
            "Referer": "https://cloud.mail.ru/"
        })
        adapter = HTTPAdapter(
            max_retries=Retry(
                total=3,
                backoff_factor=1,
                status_forcelist=[429, 500, 502, 503, 504],
                allowed_methods=["HEAD", "GET", "OPTIONS"]
            )
        )
        self.session.mount("https://", adapter)
        self.session.mount("http://", adapter)
        self.verify = verify_ssl

    def resolve_yandex(self, public_key: str):
        api = f"https://cloud-api.yandex.net/v1/disk/public/resources/download?public_key={quote(public_key, safe='')}"
        try:
            r = self.session.get(api, timeout=10, verify=self.verify)
            if r.status_code == 200:
                data = r.json()
                if "href" in data:
                    return [(data["href"], None)]
        except Exception as e:
            LOGGER.debug("Yandex resolve failed: %s", e)
        return [(public_key, None)]

    def _get_mail_base_url(self, weblink: str) -> Optional[str]:
        """Парсит base_url (weblink_get) из HTML страницы публичной папки.
        Это единственный надёжный способ для публичных папок.
        """
        try:
            public_url = f"https://cloud.mail.ru/public/{weblink.lstrip('/')}"
            r = self.session.get(public_url, timeout=15, verify=self.verify)
            if r.status_code != 200:
                return None
            
            # Метод из gist: ищем weblink_get в HTML
            for line in r.content.split(b'\n'):
                parts = line.split(b'"weblink_get":[{"count":1,"url":"')
                if len(parts) > 1:
                    base_url = parts[1].split(b'"')[0].decode('utf-8', errors='ignore')
                    return base_url.replace('\\/', '/')
            
            # Fallback: regex по всему HTML
            m = re.search(r'"weblink_get"\s*:\s*\[\s*{\s*"count"\s*:\s*1\s*,\s*"url"\s*:\s*"([^"]+)"', r.text)
            if m:
                return m.group(1).replace('\\/', '/')
                
        except Exception as e:
            LOGGER.debug("Mail base URL parse failed: %s", e)
        return None

    def _get_mail_token(self) -> Optional[str]:
        """Получает download token. Парсим JSON вручную — просто и надёжно."""
        try:
            r = self.session.get(
                "https://cloud.mail.ru/api/v2/tokens/download",
                timeout=10, verify=self.verify
            )
            if r.status_code == 200:
                # Парсим как в gist: split по кавычкам, берём 5-й элемент
                text = r.text
                parts = text.split('"')
                if len(parts) > 5:
                    return parts[5]
                # Fallback: JSON parse
                data = r.json()
                if data.get("status") == 200:
                    return data.get("body", {}).get("token") or data.get("token")
        except Exception as e:
            LOGGER.debug("Mail token failed: %s", e)
        return None

    def _list_mail_folder(self, weblink: str) -> List[Tuple[str, Optional[str]]]:
        """Достаёт ВСЕ файлы из публичной папки cloud.mail.ru.
        Алгоритм из gist: base_url из HTML → token → листинг с пагинацией.
        """
        results = []
        try:
            # 1. Базовый URL из HTML (обязательно!)
            base_url = self._get_mail_base_url(weblink)
            if not base_url:
                LOGGER.warning("Mail folder: cannot get base_url for %s", weblink)
                return []

            # 2. Токен
            token = self._get_mail_token()
            if not token:
                LOGGER.warning("Mail folder: cannot get download token")
                return []

            # 3. Чистим weblink для API
            clean = weblink.lstrip('/')
            if clean.startswith('public/'):
                clean = clean[len('public/'):]
            
            # 4. Листим с пагинацией
            offset = 0
            limit = 500
            
            while True:
                list_url = (
                    f"https://cloud.mail.ru/api/v2/folder"
                    f"?weblink={quote(clean, safe='~@#$()*!=:;,.?/\\')}"
                    f"&offset={offset}&limit={limit}&api=2"
                )
                
                r = self.session.get(list_url, timeout=20, verify=self.verify)
                if r.status_code != 200:
                    LOGGER.debug("Mail folder list HTTP %s", r.status_code)
                    break
                
                try:
                    data = r.json()
                except Exception:
                    break
                
                if data.get("status") != 200:
                    LOGGER.debug("Mail folder list status %s", data.get("status"))
                    break
                
                body = data.get("body", {})
                items = body.get("list", [])
                if not items:
                    break
                
                for item in items:
                    item_type = item.get("type") or item.get("kind", "")
                    name = item.get("name", "")
                    item_weblink = item.get("weblink", "")
                    
                    if item_type in ("folder", "dir"):
                        # Рекурсия в подпапки
                        sub_results = self._list_mail_folder(item_weblink)
                        results.extend(sub_results)
                    else:
                        # Формируем прямую ссылку на скачивание
                        # weblink файла = его путь относительно корня папки
                        file_path = item_weblink
                        if name and not file_path.endswith('/' + name):
                            file_path = file_path + '/' + name
                        
                        dl_url = f"{base_url}/{quote(file_path, safe='~@#$()*!=:;,.?/\\')}?key={token}"
                        results.append((dl_url, None))
                
                # Проверяем, есть ли ещё
                total = body.get("count", {})
                total_items = total.get("folders", 0) + total.get("files", 0)
                if offset + len(items) >= total_items:
                    break
                offset += limit
                if offset > 20000:  # защита
                    break
            
            LOGGER.info("Mail folder %s: found %d files", weblink, len(results))
            return results
            
        except Exception as e:
            LOGGER.warning("Mail folder list error: %s", e)
        return results

 
    def resolve_mail(self, public_key: str):
        parsed = urlparse(public_key)
        path = parsed.path
        weblink = path[len("/public"):] if path.startswith("/public/") else path

        # 1. Пробуем API file — самый надёжный способ для одиночного файла
        try:
            r = self.session.get(
                f"https://cloud.mail.ru/api/v2/file?weblink={quote(weblink, safe='')}",
                timeout=10, verify=self.verify
            )
            if r.status_code == 200:
                data = r.json()
                if data.get("status") == 200:
                    body = data.get("body", {})
                    # Если это папка (kind == "folder") — достаём содержимое
                    if body.get("kind") == "folder" or body.get("type") == "folder":
                        files = self._list_mail_folder(weblink)
                        if files:
                            return files
                        # Fallback на HTML-парсинг
                        files = self._extract_mail_folder_from_html(r.text, public_key)
                        if files:
                            return files
                    if body.get("download_url"):
                        return [(body["download_url"], None)]
                    wg = body.get("weblink_get", [])
                    if wg and wg[0].get("url"):
                        return [(wg[0]["url"], None)]
        except Exception as e:
            LOGGER.debug("Mail file API failed: %s", e)

        # 2. Discovery API
        try:
            r = self.session.get(
                f"https://cloud.mail.ru/api/v2/discovery?weblink={quote(weblink, safe='')}",
                timeout=10, verify=self.verify
            )
            if r.status_code == 200:
                data = r.json()
                if data.get("status") == 200:
                    body = data.get("body", {})
                    wg = body.get("weblink_get", [])
                    if wg and wg[0].get("url"):
                        # Если это папка — достаём содержимое
                        if body.get("kind") == "folder" or body.get("type") == "folder":
                            files = self._list_mail_folder(weblink)
                            if files:
                                return files
                            # Fallback на HTML-парсинг
                            files = self._extract_mail_folder_from_html(r.text, public_key)
                            if files:
                                return files
                        # Одиночный файл
                        return [(wg[0]["url"], None)]
        except Exception as e:
            LOGGER.debug("Mail discovery failed: %s", e)

        # 3. Парсим HTML — ищем прямую ссылку или список файлов в папке
        try:
            r = self.session.get(public_key, timeout=10, verify=self.verify)
            if r.status_code == 200:
                html = r.text
                # Сначала пробуем найти список файлов в папке через JSON в HTML
                files = self._extract_mail_folder_from_html(html, public_key)
                if files:
                    return files

                # Пробуем gist-метод: base_url из HTML + токен + API листинг
                folder_files = self._list_mail_folder(weblink)
                if folder_files:
                    return folder_files

                m = re.search(r'data-url="([^"]+)"', html)
                if m:
                    return [(m.group(1), None)]
                m = re.search(r'"downloadUrl":"([^"]+)"', html)
                if m:
                    return [(m.group(1).replace('\\/', '/'), None)]
                m = re.search(r'"weblink_get":\s*\[\s*{\s*"url":\s*"([^"]+)"', html)
                if m:
                    return [(m.group(1).replace('\\/', '/'), None)]
                m = re.search(r'"fileUrl":"([^"]+)"', html)
                if m:
                    return [(m.group(1).replace('\\/', '/'), None)]
                # og:image / twitter:image
                m = re.search(r'<meta[^>]+property=["\']og:image["\'][^>]+content=["\']([^"\']+)["\']', html, re.I)
                if m:
                    return [(m.group(1), None)]
        except Exception as e:
            LOGGER.debug("Mail HTML parse failed: %s", e)

        # 4. ?download=1
        dl_url = public_key
        if "?" not in dl_url:
            dl_url += "?download=1"
        else:
            dl_url += "&download=1"
        try:
            r = self.session.head(dl_url, timeout=5, verify=self.verify, allow_redirects=True)
            if r.status_code == 200:
                ct = r.headers.get("Content-Type", "").lower()
                if "html" not in ct:
                    return [(dl_url, None)]
        except Exception:
            pass

        # 5. cloclo fallback
        cloclo_url = f"https://cloclo3.cloud.mail.ru/weblink/view{weblink}"
        try:
            r = self.session.head(cloclo_url, timeout=5, verify=self.verify, allow_redirects=True)
            if r.status_code == 200:
                ct = r.headers.get("Content-Type", "").lower()
                if "image" in ct or "pdf" in ct or "octet-stream" in ct:
                    return [(cloclo_url, None)]
        except Exception:
            pass

        return [(dl_url, None)]

        # 3. Парсим HTML — ищем прямую ссылку или список файлов в папке
        try:
            r = self.session.get(public_key, timeout=10, verify=self.verify)
            if r.status_code == 200:
                html = r.text
                # Сначала пробуем найти список файлов в папке через JSON в HTML
                files = self._extract_mail_folder_from_html(html, public_key)
                if files:
                    return files
                
                # Пробуем gist-метод: base_url из HTML + токен + API листинг
                folder_files = self._list_mail_folder(weblink)
                if folder_files:
                    return folder_files

                m = re.search(r'data-url="([^"]+)"', html)
                if m:
                    return [(m.group(1), None)]
                m = re.search(r'"downloadUrl":"([^"]+)"', html)
                if m:
                    return [(m.group(1).replace('\\/', '/'), None)]
                m = re.search(r'"weblink_get":\s*\[\s*{\s*"url":\s*"([^"]+)"', html)
                if m:
                    return [(m.group(1).replace('\\/', '/'), None)]
                m = re.search(r'"fileUrl":"([^"]+)"', html)
                if m:
                    return [(m.group(1).replace('\\/', '/'), None)]
                # og:image / twitter:image
                m = re.search(r'<meta[^>]+property=["\']og:image["\'][^>]+content=["\']([^"\']+)["\']', html, re.I)
                if m:
                    return [(m.group(1), None)]
        except Exception as e:
            LOGGER.debug("Mail HTML parse failed: %s", e)

        # 4. ?download=1
        dl_url = public_key
        if "?" not in dl_url:
            dl_url += "?download=1"
        else:
            dl_url += "&download=1"
        try:
            r = self.session.head(dl_url, timeout=5, verify=self.verify, allow_redirects=True)
            if r.status_code == 200:
                ct = r.headers.get("Content-Type", "").lower()
                if "html" not in ct:
                    return [(dl_url, None)]
        except Exception:
            pass

        # 5. cloclo fallback
        cloclo_url = f"https://cloclo3.cloud.mail.ru/weblink/view{weblink}"
        try:
            r = self.session.head(cloclo_url, timeout=5, verify=self.verify, allow_redirects=True)
            if r.status_code == 200:
                ct = r.headers.get("Content-Type", "").lower()
                if "image" in ct or "pdf" in ct or "octet-stream" in ct:
                    return [(cloclo_url, None)]
        except Exception:
            pass

        return [(dl_url, None)]

    def _extract_mail_folder_from_html(self, html: str, base_url: str) -> List[Tuple[str, Optional[str]]]:
        """Парсит HTML-страницу папки mail.ru — ищет ссылки на отдельные файлы."""
        results = []
        # Ищем все ссылки на /public/ в href — это могут быть файлы/подпапки
        for m in re.finditer(r'href="(https?://cloud\.mail\.ru/public/[A-Za-z0-9/_\-]+)"', html):
            sub_url = m.group(1)
            # Извлекаем имя файла (последний компонент пути)
            parts = sub_url.rstrip('/').split('/')
            name = parts[-1] if parts else None
            if name and name not in ['public']:
                results.append((sub_url, name))
        if results:
            return results
        # Альтернатива: JSON ссылки внутри HTML
        for m in re.finditer(r'"(?:weblink|file|url|src|downloadUrl)"\s*:\s*"([^"]+/public/[A-Za-z0-9/_\-]+)"', html, re.I):
            u = m.group(1).replace('\\/', '/')
            name = u.rstrip('/').split('/')[-1]
            results.append((u, name))
        return results

    # ============================================================
    # GOOGLE DRIVE
    # ============================================================
    @staticmethod
    def parse_google_id(url: str) -> Optional[str]:
        """Извлекает file/folder ID из любой формы ссылки Google Drive."""
        if not url:
            return None
        # open?id=ID / uc?id=ID / uc?export=download&id=ID / thumbnail?id=ID
        m = re.search(r'[?&]id=([a-zA-Z0-9_\-]+)', url)
        if m:
            return m.group(1)
        # /file/d/ID/ или /folders/ID или /drive/folders/ID
        m = re.search(r'/(?:file/d/|folders/|drive/folders/|drive/u/\d+/folders/)([a-zA-Z0-9_\-]+)', url)
        if m:
            return m.group(1)
        return None

    @staticmethod
    def is_google_drive_folder(url: str) -> bool:
        if not url:
            return False
        u = url.lower()
        return ('/folders/' in u or '/folderview' in u or 'type=folder' in u) and 'drive.google.com' in u

    def resolve_google_drive(self, url: str):
        """Возвращает список (download_url, subfolder) для Google Drive.
        Если ссылка на папку — пытается достать список файлов внутри неё.
        """
        file_id = self.parse_google_id(url)
        if not file_id:
            return [(url, None)]

        # Если это папка — пытаемся собрать все файлы из неё
        if self.is_google_drive_folder(url):
            files = self._list_google_drive_folder(file_id)
            if files:
                return files
            # Если папку открыть не удалось — возвращаем как есть
            return [(url, None)]

        # Для одиночного файла пробуем несколько стратегий
        candidates = [
            f"https://drive.google.com/uc?export=download&id={file_id}",
            f"https://drive.google.com/uc?export=view&id={file_id}",
            f"https://drive.google.com/thumbnail?id={file_id}&sz=w2000",
            f"https://lh3.googleusercontent.com/d/{file_id}=s0",
            f"https://lh3.googleusercontent.com/d/{file_id}=w2000",
        ]
        results = []
        for cu in candidates:
            try:
                r = self.session.get(
                    cu, timeout=15, verify=self.verify, allow_redirects=True,
                    headers={"Referer": "https://drive.google.com/"}
                )
                if r.status_code != 200:
                    continue
                ct = r.headers.get("Content-Type", "").lower()
                if "text/html" in ct:
                    # Для больших файлов Google показывает страницу-предупреждение
                    # с confirm-токеном — пробуем его извлечь
                    html = r.text
                    m = re.search(r'href="(/uc\?export=download[^"]+)"', html)
                    if m:
                        confirm_url = "https://drive.google.com" + m.group(1).replace("&amp;", "&")
                        results.append((confirm_url, None))
                        continue
                    m = re.search(r'"downloadUrl"\s*:\s*"([^"]+)"', html)
                    if m:
                        results.append((m.group(1).replace("\\u003d", "=").replace("\\/", "/"), None))
                        continue
                    # og:image / og:video — fallback
                    m = re.search(r'<meta[^>]+property=["\']og:image["\'][^>]+content=["\']([^"\']+)["\']', html, re.I)
                    if m:
                        results.append((m.group(1), None))
                        continue
                    continue
                # Бинарь получен — URL рабочий, держим первым
                # (мы не качаем сюда, только кладём URL в очередь)
                results.insert(0, (cu, None))
                # Если удалось с первого URL получить бинарь — остальные не нужны
                return results[:3]
            except Exception as e:
                LOGGER.debug("GDrive %s failed: %s", cu, e)
                continue

        return results if results else [(url, None)]

    def _list_google_drive_folder(self, folder_id: str):
        """Пытается вытащить список файлов из публичной папки Google Drive.
        Возвращает список [(url, subfolder), ...] либо [].
        """
        urls = []
        try:
            r = self.session.get(
                f"https://drive.google.com/drive/folders/{folder_id}",
                timeout=20, verify=self.verify, allow_redirects=True,
                headers={"Referer": "https://drive.google.com/"}
            )
            if r.status_code != 200:
                return []
            html = r.text
            # Паттерн 1: data-id="FILE_ID" с именем файла
            # AF_initDataCallback — JSON внутри страницы
            m = re.search(r'AF_initDataCallback\(\{key:\s*[\'"](\w+)[\'"].*?data:(.*?)\}\);', html, re.DOTALL)
            if m:
                # Грубый поиск всех ID файлов вида [null,null,"1AbCd..."]
                ids = re.findall(r'"([a-zA-Z0-9_\-]{20,})"', m.group(2))
                # Убираем дубликаты, оставляя те, что похожи на Drive ID (>=20 chars)
                seen = set()
                for fid in ids:
                    if fid in seen or fid == folder_id:
                        continue
                    seen.add(fid)
                    urls.append((
                        f"https://drive.google.com/uc?export=download&id={fid}",
                        None
                    ))
                    if len(urls) >= 50:
                        break
            if urls:
                return urls

            # Паттерн 2: og:url / прямые ссылки /file/d/
            for m in re.finditer(r'href="(/file/d/([a-zA-Z0-9_\-]+)/view[^"]*)"', html):
                fid = m.group(2)
                urls.append((
                    f"https://drive.google.com/uc?export=download&id={fid}",
                    None
                ))
                if len(urls) >= 50:
                    break
            if urls:
                return urls

            # Паттерн 3: data-id атрибуты на ссылках
            for m in re.finditer(r'data-id="([a-zA-Z0-9_\-]+)"', html):
                fid = m.group(1)
                if fid == folder_id:
                    continue
                urls.append((
                    f"https://drive.google.com/uc?export=download&id={fid}",
                    None
                ))
                if len(urls) >= 50:
                    break
        except Exception as e:
            LOGGER.debug("GDrive folder list failed: %s", e)
        return urls

    # ============================================================
    # DROPBOX
    # ============================================================
    @staticmethod
    def is_dropbox(url: str) -> bool:
        if not url:
            return False
        u = url.lower()
        return ('dropbox.com' in u) or ('dl.dropboxusercontent.com' in u)

    def resolve_dropbox(self, url: str):
        """Dropbox: меняет dl=0 на dl=1, raw=1 для прямого скачивания."""
        try:
            # Ссылка уже прямая (dl.dropboxusercontent.com) — оставляем
            if 'dl.dropboxusercontent.com' in url:
                return [(url, None)]
            # dl=0 -> dl=1, иначе добавляем ?dl=1
            if 'dl=' in url:
                new_url = re.sub(r'dl=\d', 'dl=1', url)
            elif '?' in url:
                new_url = url + '&dl=1'
            else:
                new_url = url + '?dl=1'
            return [(new_url, None)]
        except Exception:
            return [(url, None)]

    # ============================================================
    # ONEDRIVE / SHAREPOINT
    # ============================================================
    @staticmethod
    def is_onedrive(url: str) -> bool:
        if not url:
            return False
        u = url.lower()
        return any(d in u for d in [
            'onedrive.live.com', '1drv.ms', 'sharepoint.com',
            'office.com', 'skydrive.com', 'storage.live.com'
        ])

    def resolve_onedrive(self, url: str):
        """OneDrive/SharePoint — пытаемся получить embed URL или прямую ссылку."""
        try:
            # 1drv.ms — короткая ссылка, пусть редиректит сама
            if '1drv.ms' in url.lower():
                # Сначала HEAD чтобы получить финальный URL
                try:
                    r = self.session.get(
                        url, timeout=10, verify=self.verify, allow_redirects=True,
                        headers={"User-Agent": "Mozilla/5.0"}
                    )
                    final = r.url
                    return [(final, None)]
                except Exception:
                    return [(url, None)]
            # onedrive.live.com — меняем ?view на ?download
            if 'onedrive.live.com' in url.lower():
                # Попробуем добавить/download=1 или embed
                candidates = []
                if '?' in url:
                    base = url
                    candidates.append(base + ('&' if '?' in base else '?') + 'download=1')
                else:
                    candidates.append(url + '?download=1')
                return [(c, None) for c in candidates]
            return [(url, None)]
        except Exception:
            return [(url, None)]

    # ============================================================
    # HTML PARSER — поиск прямой ссылки на файл в HTML-обёртке
    # ============================================================
    def find_file_in_html(self, html: str, base_url: str) -> List[Tuple[str, Optional[str]]]:
        """Ищет прямую ссылку на файл (jpg/png/webp/pdf) внутри HTML-страницы.
        Возвращает список (url, subfolder), отсортированный по приоритету.
        """
        if not html or len(html) < 100:
            return []
        results = []

        def absolutize(u: str) -> str:
            if not u:
                return u
            if u.startswith(('http://', 'https://')):
                return u
            if u.startswith('//'):
                scheme = urlparse(base_url).scheme or 'https'
                return f"{scheme}:{u}"
            return urljoin(base_url, u)

        def score(u: str, weight: int = 0) -> int:
            """Чем выше — тем приоритетнее."""
            ul = u.lower()
            s = weight
            # Расширения файлов — самый сильный сигнал
            for ext, w in [('.jpg', 200), ('.jpeg', 200), ('.png', 200),
                          ('.webp', 200), ('.pdf', 180), ('.gif', 150), ('.bmp', 150)]:
                if ext in ul:
                    s += w
                    break
            # Google Drive / известные CDN
            if 'googleusercontent.com' in ul or 'drive.google.com' in ul:
                s += 80
            if 'yadi.sk' in ul or 'disk.yandex' in ul:
                s += 80
            if 'mail.ru' in ul:
                s += 80
            # og:image / twitter:image уже учтены в weight
            return s

        # 1. <meta http-equiv="refresh"> — redirect на другую страницу
        m = re.search(r'<meta[^>]+http-equiv=["\']refresh["\'][^>]+content=["\']\s*\d+\s*;\s*url=([^"\']+)["\']',
                      html, re.I)
        if m:
            ru = absolutize(m.group(1).replace('&amp;', '&'))
            results.append((ru, 1000))  # самый высокий приоритет — редирект

        # 2. og:image / og:image:secure_url / twitter:image
        for pat, w in [
            (r'<meta[^>]+property=["\']og:image:secure_url["\'][^>]+content=["\']([^"\']+)["\']', 950),
            (r'<meta[^>]+property=["\']og:image["\'][^>]+content=["\']([^"\']+)["\']', 940),
            (r'<meta[^>]+name=["\']twitter:image["\'][^>]+content=["\']([^"\']+)["\']', 935),
            (r'<link[^>]+rel=["\']image_src["\'][^>]+href=["\']([^"\']+)["\']', 930),
        ]:
            for mm in re.finditer(pat, html, re.I):
                results.append((absolutize(mm.group(1)), w))

        # 3. Прямые ссылки на файлы в href/src
        for mm in re.finditer(r'(?:href|src)\s*=\s*["\']([^"\']+\.(?:jpg|jpeg|png|webp|pdf|gif|bmp)(?:\?[^"\']*)?)["\']',
                              html, re.I):
            u = absolutize(mm.group(1))
            if 'sprite' in u.lower() or 'icon' in u.lower() or 'logo' in u.lower():
                continue
            results.append((u, score(u, 700)))

        # 4. JSON-LD или встроенный JSON с ссылкой
        for mm in re.finditer(r'"(?:image|contentUrl|downloadUrl|fileUrl|src)"\s*:\s*"([^"]+)"', html, re.I):
            u = absolutize(mm.group(1).replace('\\/', '/').replace('\\u0026', '&'))
            if any(ext in u.lower() for ext in ['.jpg', '.jpeg', '.png', '.webp', '.pdf']):
                results.append((u, score(u, 600)))

        # 5. Кнопка "Скачать" / data-атрибуты
        for mm in re.finditer(r'data-(?:url|src|file|download|url)\s*=\s*"([^"]+)"', html, re.I):
            u = absolutize(mm.group(1))
            results.append((u, score(u, 500)))

        # 6. Большие <img> (по атрибутам width/height или src без мелких значков)
        for mm in re.finditer(r'<img[^>]+src=["\']([^"\']+)["\']', html, re.I):
            u = absolutize(mm.group(1))
            if 'sprite' in u.lower() or 'icon' in u.lower() or 'avatar' in u.lower():
                continue
            results.append((u, score(u, 300)))

        # 7. data-src (lazy load)
        for mm in re.finditer(r'data-src=["\']([^"\']+)["\']', html, re.I):
            u = absolutize(mm.group(1))
            if 'sprite' in u.lower() or 'icon' in u.lower():
                continue
            results.append((u, score(u, 290)))

        # Сортируем по score и убираем дубликаты
        seen = set()
        unique = []
        for url, sc in sorted(results, key=lambda x: -x[1]):
            if url in seen:
                continue
            seen.add(url)
            unique.append(url)
        return [(u, None) for u in unique]

    # ============================================================
    # ROUTER — главная точка входа для распознавания URL
    # ============================================================
    def resolve(self, url: str, aggressive: bool = True):
        """Универсальный resolver: возвращает [(url, subfolder), ...]
        для любого поддерживаемого облака. Если ничего не подошло — [(url, None)].

        aggressive=False — для прямых ссылок: папки не раскрываются, HTML-парсинг не делается.
        """
        try:
            ul = url.lower()
            if 'drive.google.com' in ul:
                if aggressive:
                    return self.resolve_google_drive(url)
                else:
                    # Только прямая ссылка на файл, без попыток открыть папку
                    fid = self.parse_google_id(url)
                    if fid and not self.is_google_drive_folder(url):
                        return [(f"https://drive.google.com/uc?export=download&id={fid}", None)]
                    return [(url, None)]
            if self.is_dropbox(url):
                return self.resolve_dropbox(url)
            if self.is_onedrive(url):
                return self.resolve_onedrive(url)
            if 'yadi.sk' in ul or 'disk.yandex' in ul or '360.yandex' in ul:
                return self.resolve_yandex(url)
            if 'cloud.mail.ru' in ul:
                if aggressive:
                    return self.resolve_mail(url)
                else:
                    # Только ?download=1 — без попыток открыть папку
                    if '?' in url:
                        return [(url + '&download=1', None)]
                    return [(url + '?download=1', None)]
        except Exception as e:
            LOGGER.debug("resolve(%s) failed: %s", url, e)
        return [(url, None)]


# ============================================================
# CRASH REPORTER — пишет детальный отчёт в текстовый файл
# ============================================================
def write_crash_report(exc_info=None, context: str = ""):
    """Записывает подробный crash-отчёт в текстовый файл рядом со скриптом.
    Используется при любом необработанном исключении или ручном вызове.
    """
    try:
        # Куда писать: рядом со скриптом, или в домашнюю папку если она недоступна
        try:
            base_dir = os.path.dirname(os.path.abspath(sys.argv[0]))
            test_path = os.path.join(base_dir, "_test_write.tmp")
            with open(test_path, "w") as f:
                f.write("test")
            os.remove(test_path)
            report_dir = base_dir
        except Exception:
            report_dir = os.path.expanduser("~")

        ts = time.strftime("%Y%m%d_%H%M%S")
        report_path = os.path.join(report_dir, f"valera_crash_{ts}.txt")

        lines = []
        lines.append("=" * 70)
        lines.append("VALERA CRASH REPORT")
        lines.append("=" * 70)
        lines.append(f"Время:           {time.strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append(f"Контекст:        {context or '(не указан)'}")
        lines.append(f"Python:          {sys.version}")
        lines.append(f"Платформа:       {sys.platform}")
        try:
            import platform
            lines.append(f"ОС:              {platform.platform()}")
            lines.append(f"Архитектура:     {platform.machine()}")
        except Exception:
            pass
        try:
            from PySide6 import __version__ as qt_ver
            lines.append(f"PySide6:         {qt_ver}")
        except Exception:
            pass

        lines.append("")
        lines.append("-" * 70)
        lines.append("ТЕКУЩАЯ РАБОТА")
        lines.append("-" * 70)
        try:
            cwd = os.getcwd()
            lines.append(f"cwd:             {cwd}")
        except Exception:
            pass

        lines.append("")
        lines.append("-" * 70)
        lines.append("TRACEBACK")
        lines.append("-" * 70)
        if exc_info is not None:
            t, v, tb = exc_info
            lines.append("".join(traceback.format_exception(t, v, tb)))
        else:
            # Если exc_info не передан — возьмём текущий
            lines.append(traceback.format_exc())

        lines.append("")
        lines.append("-" * 70)
        lines.append("ЗАВЕРШЕНО")
        lines.append("-" * 70)

        with open(report_path, "w", encoding="utf-8") as f:
            f.write("\n".join(lines))

        LOGGER.critical("Crash report written to: %s", report_path)
        return report_path
    except Exception as e:
        # Если даже crash report не удалось записать — просто лог
        LOGGER.critical("Failed to write crash report: %s", e)
        return None


# ============================================================
# WORKER
# ============================================================
class Worker(QThread):
    progress = Signal(int)
    stats_updated = Signal(int, int, int, int, int, float)
    log_row = Signal(str, str, str)
    finished = Signal(dict)
    error = Signal(str)
    download_errors_ready = Signal(list)

    def __init__(self, settings: AppSettings):
        super().__init__()
        self.s = settings
        self._p = False
        self._c = False
        self._m = QMutex()
        self._w = QWaitCondition()
        self.stats = {"ok": 0, "defect_size": 0, "defect_white": 0, "fail": 0, "bytes": 0}
        self.processor = ImageProcessor()
        self.resolver = CloudResolver(verify_ssl=settings.ssl_verify)
        self.download_errors = []
        self.start_time = 0
        self.executor = None  # ThreadPoolExecutor — хранится здесь, чтобы можно было корректно отменить
        self.session_closed = False
        # Диагностика submit-фазы: кол-во поставленных в очередь и завершённых futures
        self._dl_submitted = 0
        self._dl_done = 0
        self._dl_total = 0  # плановое количество futures (обновляется по ходу)
        self._last_progress = -1  # последнее emit'нутое значение прогресса
        self._last_log_time = 0   # для throttle логов в submit-фазе

    def pause(self):
        self._p = True

    def resume(self):
        self._p = False
        self._w.wakeAll()

    def cancel(self):
        """Корректная отмена: прерывает загрузки и завершает воркер без висящих процессов."""
        self._c = True
        self.resume()

        # 1. Закрываем HTTP session — это прервёт все текущие .get() с исключением
        try:
            if not self.session_closed:
                self.resolver.session.close()
                self.session_closed = True
        except Exception as e:
            LOGGER.debug("Session close on cancel failed: %s", e)

        # 2. Отменяем ThreadPoolExecutor
        # cancel_futures=True отменяет ещё не запущенные futures
        # wait=False — не ждём завершения запущенных (они прервутся через закрытую session)
        if self.executor is not None:
            try:
                self.executor.shutdown(wait=False, cancel_futures=True)
            except Exception as e:
                LOGGER.debug("Executor shutdown failed: %s", e)

    def check_cancel(self) -> bool:
        return self._c

    def check_pause(self):
        if self._p:
            self._m.lock()
            self._w.wait(self._m)
            self._m.unlock()

    def emit_stats(self):
        elapsed = time.time() - self.start_time
        self.stats_updated.emit(
            self.stats["ok"], self.stats["defect_size"],
            self.stats["defect_white"], self.stats["fail"],
            self.stats["bytes"], elapsed
        )

    def log(self, status: str, source: str, message: str = ""):
        self.log_row.emit(status, source, message)

    def save_img(self, img: PILImage.Image, path: str):
        fmt = self.s.fmt
        try:
            if fmt == "jpg":
                if img.mode in ("RGBA", "P"):
                    rgb = ensure_rgb(img)
                    rgb.save(path, quality=Defaults.JPEG_QUALITY, optimize=True)
                    rgb.close()
                else:
                    img.save(path, quality=Defaults.JPEG_QUALITY, optimize=True)
            elif fmt == "webp":
                img.save(path, optimize=True, quality=Defaults.WEBP_QUALITY)
            else:
                img.save(path, optimize=True)
        finally:
            try:
                img.close()
            except Exception:
                pass

    def process_pdf_bytes(self, content: bytes):
        try:
            import fitz
        except ImportError:
            raise Exception("pip install PyMuPDF")
        doc = fitz.open(stream=content, filetype="pdf")
        imgs = []
        for pn in range(len(doc)):
            pg = doc.load_page(pn)
            px = pg.get_pixmap(dpi=Defaults.PDF_DPI)
            imgs.append((
                PILImage.frombytes("RGB", [px.width, px.height], px.samples).convert("RGBA"),
                pn + 1
            ))
        doc.close()
        return imgs

    def is_pdf_multi_b(self, c: bytes) -> bool:
        try:
            import fitz
            doc = fitz.open(stream=c, filetype="pdf")
            l = len(doc)
            doc.close()
            return l > 1
        except Exception:
            return False

    def is_pdf_multi_f(self, fp: str) -> bool:
        try:
            import fitz
            doc = fitz.open(fp)
            l = len(doc)
            doc.close()
            return l > 1
        except Exception:
            return False

    def run(self):
        try:
            self.start_time = time.time()
            self.stats = {"ok": 0, "defect_size": 0, "defect_white": 0, "fail": 0, "bytes": 0}
            self.download_errors.clear()
            if self.s.source_type == "Отбракованное":
                self.process_rejected()
            elif self.s.source_type == "Excel":
                self.process_excel()
            else:
                self.process_folder(self.s.source, self.s.out_dir)
            self.stats["time"] = time.time() - self.start_time
            self.download_errors_ready.emit(self.download_errors)
            self.finished.emit(self.stats)
        except Exception:
            LOGGER.exception("Worker crash")
            # Пишем crash-отчёт в файл рядом со скриптом
            ctx = f"source_type={self.s.source_type}, source={self.s.source}, out_dir={self.s.out_dir}"
            write_crash_report(context=ctx)
            self.error.emit(traceback.format_exc())
        finally:
            # Всегда закрываем session и executor — даже если было исключение или отмена
            try:
                if self.executor is not None and not self._c:
                    self.executor.shutdown(wait=True)
            except Exception:
                pass
            try:
                if not self.session_closed:
                    self.resolver.session.close()
                    self.session_closed = True
            except Exception:
                pass

    def process_rejected(self):
        files = self.s.selected_rejected_files
        if files:
            self.stats["processed_dir"] = os.path.dirname(files[0])
        if not files:
            return
        thr = self.s.white_threshold
        for i, fp in enumerate(files):
            if self.check_cancel():
                break
            self.check_pause()
            self._proc_single(fp, thr)
            self.progress.emit(int((i + 1) / len(files) * 100))
        self.emit_stats()
        self.progress.emit(100)

    def _proc_single(self, fp: str, threshold: int = Defaults.THRESHOLD_WHITE):
        bn = os.path.basename(fp)
        ft = ""
        if bn.startswith("!РАЗМЕР_"):
            ft = "SIZE"
        elif bn.startswith("!БЕЛЫЙ_"):
            ft = "WHITE_BG"
        elif bn.startswith("!ОШИБКА_"):
            ft = "RETRY"
        if not ft:
            return
        cn = bn
        for p in ["!РАЗМЕР_", "!БЕЛЫЙ_", "!ОШИБКА_"]:
            if cn.startswith(p):
                cn = cn[len(p):]
                break
        fb = os.path.splitext(cn)[0]
        td = os.path.dirname(fp)
        tp = os.path.join(td, f"{fb}_ИСПРАВЛЕНО.{self.s.fmt}")
        c = 1
        while os.path.exists(tp):
            tp = os.path.join(td, f"{fb}_ИСПРАВЛЕНО_{c}.{self.s.fmt}")
            c += 1
        try:
            with PILImage.open(fp) as img:
                img = img.convert("RGBA")
                pi, _, _ = self.processor.process_with_analysis(img, self.s, force_type=ft, threshold=threshold)
            self.save_img(pi, tp)
            self.stats["ok"] += 1
        except Exception as e:
            self.stats["fail"] += 1
            self.log("[ОШИБКА]", bn, str(e))
        self.emit_stats()

    def process_folder(self, src: str, out: str):
        dd = os.path.join(out, "Готово")
        self.stats["processed_dir"] = dd
        files = [
            os.path.join(r, f)
            for r, _, fs in os.walk(src)
            for f in fs
            if is_processable_file(f)
        ]
        thr = self.s.white_threshold
        for i, fp in enumerate(files):
            if self.check_cancel():
                break
            self.check_pause()
            try:
                rp = os.path.relpath(fp, src)
                tfp = os.path.join(dd, rp)
                os.makedirs(os.path.dirname(tfp), exist_ok=True)
                fb = os.path.splitext(os.path.basename(fp))[0]
                if fp.lower().endswith(".pdf"):
                    if self.is_pdf_multi_f(fp):
                        shutil.copy2(fp, os.path.join(os.path.dirname(tfp), f"{fb}.pdf"))
                        self.stats["ok"] += 1
                    else:
                        self._proc_pdf(fp, tfp, fb, thr)
                else:
                    self._proc_img(fp, tfp, fb, thr)
            except Exception:
                self.stats["fail"] += 1
            self.emit_stats()
            self.progress.emit(int((i + 1) / max(1, len(files)) * 100))

    def _proc_pdf(self, fp: str, tfp: str, fb: str, threshold: int = Defaults.THRESHOLD_WHITE):
        with open(fp, "rb") as f:
            c = f.read()
        for img, pn in self.process_pdf_bytes(c):
            if self.check_cancel():
                img.close()
                break
            force = "PDF_SQUARE" if self.s.pdf_always_square else ""
            pi, dp, cat = self.processor.process_with_analysis(img, self.s, force_type=force, threshold=threshold)
            img.close()
            self.save_img(pi, os.path.join(os.path.dirname(tfp), f"{dp}{fb}_стр{pn}.{self.s.fmt}"))
            self._log_def(fb + f"_стр{pn}", cat)

    def _proc_img(self, fp: str, tfp: str, fb: str, threshold: int = Defaults.THRESHOLD_WHITE):
        with PILImage.open(fp) as img:
            img = img.convert("RGBA")
            pi, dp, cat = self.processor.process_with_analysis(img, self.s, threshold=threshold)
        self.save_img(pi, os.path.join(os.path.dirname(tfp), f"{dp}{fb}.{self.s.fmt}"))
        self._log_def(fb, cat)
        self.stats["bytes"] += os.path.getsize(fp)

    def _log_def(self, n: str, cat: str):
        if cat == "SIZE_FAIL":
            self.stats["defect_size"] += 1
            self.log("[ОТБРАКОВАНО]", n, "!РАЗМЕР_")
        elif cat == "QUESTION_WHITE":
            self.stats["defect_white"] += 1
            self.log("[ПОД ВОПРОСОМ]", n, "!БЕЛЫЙ_")
        else:
            self.stats["ok"] += 1
            self.log("[OK]", n, "Успешно")

    def process_excel(self):
        eb = os.path.splitext(os.path.basename(self.s.source))[0]
        dl_dir = os.path.join(self.s.out_dir, f"скачано_{eb}")
        pr_dir = os.path.join(self.s.out_dir, f"обработано_{eb}")
        os.makedirs(dl_dir, exist_ok=True)
        os.makedirs(pr_dir, exist_ok=True)
        self.stats["processed_dir"] = pr_dir
        xl = ExcelManager(self.s.source)
        wb = xl.load()
        fc = col_letter_to_index(self.s.url_from)
        tc = col_letter_to_index(self.s.url_to)
        ucols = list(range(fc, tc + 1))
        ac = col_letter_to_index(self.s.article_col)
        mc = max(ac, max(ucols)) + 1
        tu = xl.count_urls(ucols, ac)
        if tu == 0:
            xl.close()
            raise Exception("В Excel нет ссылок!")
        cm = {}
        proc = 0
        # Прогресс для скачивания: 0-70% (а не 50%, чтобы обработка имела свой запас)
        DL_PROGRESS_MAX = 70

        def download_task(fetch_url: str, raw_folder: str, idx: int):
            os.makedirs(raw_folder, exist_ok=True)
            try:
                resp = self.resolver.session.get(
                    fetch_url, timeout=120,
                    headers={"Referer": f"{urlparse(fetch_url).scheme}://{urlparse(fetch_url).netloc}/"},
                    verify=self.s.ssl_verify, allow_redirects=True
                )
                if resp.status_code != 200:
                    raise Exception(f"Status {resp.status_code}")
                content = resp.content
                ctype = resp.headers.get("Content-Type", "").lower()
                cd = resp.headers.get("Content-Disposition", "").lower()

                # Если получили HTML — пытаемся найти прямую ссылку на файл
                if "text/html" in ctype and len(content) > 1024:
                    html = content[:65536].decode('utf-8', errors='ignore')

                    # 1. <meta http-equiv="refresh"> — короткий редирект
                    m = re.search(r'<meta[^>]+http-equiv=["\']refresh["\'][^>]+content=["\']\s*\d+\s*;\s*url=([^"\']+)["\']',
                                  html, re.I)
                    if m:
                        redirect_url = m.group(1).replace('&amp;', '&')
                        if not redirect_url.startswith(('http://', 'https://')):
                            redirect_url = urljoin(fetch_url, redirect_url)
                        try:
                            resp2 = self.resolver.session.get(
                                redirect_url, timeout=60,
                                headers={"Referer": fetch_url},
                                verify=self.s.ssl_verify, allow_redirects=True
                            )
                            if resp2.status_code == 200 and \
                               "text/html" not in resp2.headers.get("Content-Type", "").lower():
                                content = resp2.content
                                ctype = resp2.headers.get("Content-Type", "").lower()
                                cd = resp2.headers.get("Content-Disposition", "").lower()
                        except Exception as e:
                            LOGGER.debug("HTML meta-refresh follow failed: %s", e)

                    # 2. Поиск прямой ссылки на файл внутри HTML-обёртки
                    if "text/html" in ctype and self.s.aggressive_parse:
                        candidates = self.resolver.find_file_in_html(html, fetch_url)
                        for cand_url, _ in candidates[:5]:  # пробуем первые 5 вариантов
                            try:
                                resp3 = self.resolver.session.get(
                                    cand_url, timeout=60,
                                    headers={"Referer": fetch_url},
                                    verify=self.s.ssl_verify, allow_redirects=True
                                )
                                if resp3.status_code != 200:
                                    continue
                                ct3 = resp3.headers.get("Content-Type", "").lower()
                                if "text/html" in ct3:
                                    continue
                                # Нашли бинарь — используем его
                                content = resp3.content
                                ctype = ct3
                                cd = resp3.headers.get("Content-Disposition", "").lower()
                                fetch_url = cand_url  # для логов
                                break
                            except Exception as e:
                                LOGGER.debug("HTML candidate %s failed: %s", cand_url, e)
                                continue

                    # Если так и остались HTML — последний шанс: снова meta-refresh
                    if "text/html" in ctype:
                        m = re.search(r'content=["\']\s*\d+\s*;\s*url=([^"\']+)["\']', html, re.I)
                        if m and self.s.aggressive_parse:
                            redirect_url = m.group(1).replace('&amp;', '&')
                            if not redirect_url.startswith(('http://', 'https://')):
                                redirect_url = urljoin(fetch_url, redirect_url)
                            try:
                                resp4 = self.resolver.session.get(
                                    redirect_url, timeout=60,
                                    headers={"Referer": fetch_url},
                                    verify=self.s.ssl_verify, allow_redirects=True
                                )
                                if resp4.status_code == 200 and \
                                   "text/html" not in resp4.headers.get("Content-Type", "").lower():
                                    content = resp4.content
                                    ctype = resp4.headers.get("Content-Type", "").lower()
                                    cd = resp4.headers.get("Content-Disposition", "").lower()
                                    fetch_url = redirect_url
                            except Exception:
                                pass

                    if "text/html" in ctype:
                        raise Exception("Получена HTML-страница вместо файла. Ссылка устарела или требует авторизации.")

                is_pdf = (
                    fetch_url.lower().endswith(".pdf") or
                    "application/pdf" in ctype or
                    ".pdf" in cd or
                    (len(content) >= 5 and content[:5] == b'%PDF-')
                )
                if is_pdf:
                    if self.is_pdf_multi_b(content):
                        pdf_path = os.path.join(raw_folder, f"raw_{idx}.pdf")
                        with open(pdf_path, "wb") as pf:
                            pf.write(content)
                        return ("OK_PDF", os.path.relpath(pdf_path, dl_dir), None, len(content))
                    else:
                        paths = []
                        for img, pn in self.process_pdf_bytes(content):
                            fname = f"raw_{idx}_{pn}.{self.s.fmt}"
                            rp = os.path.join(raw_folder, fname)
                            self.save_img(img, rp)
                            paths.append(os.path.relpath(rp, dl_dir))
                        return ("OK_PDF_IMG", paths, None, len(content))
                else:
                    ext = ".jpg"
                    if "image/png" in ctype or fetch_url.lower().endswith(".png"):
                        ext = ".png"
                    elif "image/webp" in ctype or fetch_url.lower().endswith(".webp"):
                        ext = ".webp"
                    detected = detect_ext_from_bytes(content)
                    if detected:
                        ext = detected
                    fname = f"raw_{idx}{ext}"
                    rp = os.path.join(raw_folder, fname)
                    with open(rp, "wb") as f:
                        f.write(content)
                    return ("OK_IMG", os.path.relpath(rp, dl_dir), None, len(content))
            except Exception as e:
                return ("ERR", None, str(e), 0)

        # Сохраняем executor в self, чтобы cancel() мог его правильно закрыть
        self.executor = ThreadPoolExecutor(max_workers=Defaults.DOWNLOAD_WORKERS)
        # Сбрасываем диагностику
        self._dl_submitted = 0
        self._dl_done = 0
        self._dl_total = 0
        self._last_progress = -1
        self._last_log_time = 0
        try:
            futures = {}

            # Эмитим сразу что начали — чтобы UI отреагировал
            self.progress.emit(1)
            self.log("[ИНФО]", "Загрузка", f"Найдено {tu} ссылок, готовлю задачи…")

            for sheet in wb.worksheets:
                ss = sanitize_filename(sheet.title)
                for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, max_col=mc):
                    if self.check_cancel():
                        break
                    if ac >= len(row) or not row[ac].value:
                        continue
                    art = row[ac].value
                    er = row[0].row
                    sa = sanitize_filename(str(art))
                    fn = f"{er:03d} - {sa}" if self.s.folder_sort == "В порядке Excel" else sa
                    tb = os.path.join(dl_dir, ss)
                    idx = 1
                    for col in ucols:
                        if col >= len(row) or not row[col].value:
                            continue
                        cell = row[col]
                        urls = extract_urls(cell.value)
                        for val in urls:
                            # Универсальный резолвер: сам поймёт, какой это облако
                            utf = self.resolver.resolve(val, aggressive=self.s.aggressive_parse)
                            for fu, sf in utf:
                                rf = os.path.join(tb, fn)
                                if sf:
                                    rf = os.path.join(rf, sf)
                                future = self.executor.submit(download_task, fu, rf, idx)
                                futures[future] = (cell, val, fu, fn)
                                idx += 1
                                self._dl_submitted += 1
                                self._dl_total = max(self._dl_total, self._dl_submitted + len(futures) - self._dl_submitted)

                                # Лог submit-фазы каждые ~5 сек или каждые 20 submit'ов
                                now = time.time()
                                if (now - self._last_log_time) > 5.0 or self._dl_submitted % 25 == 1:
                                    self.log("[ИНФО]", "Загрузка",
                                             f"Подготовлено {self._dl_submitted} задач…")
                                    self._last_log_time = now
                                    QApplication.processEvents()

                                # Progress submit-фазы: 1-5% (подготовка не считается скачиванием)
                                submit_pct = min(5, max(1, int(self._dl_submitted * 5 / max(1, tu))))
                                if submit_pct != self._last_progress:
                                    self.progress.emit(submit_pct)
                                    self._last_progress = submit_pct
                                    QApplication.processEvents()

            # Все submit'ы готовы — сообщим и продолжим
            self.log("[ИНФО]", "Загрузка",
                     f"Все {self._dl_submitted} задач в работе, ждём скачивания…")
            QApplication.processEvents()

            # Основной цикл обработки завершённых futures
            for future in as_completed(futures):
                if self.check_cancel():
                    break
                cell, val, fu, folder_name = futures[future]
                try:
                    res_type, path, err, size = future.result()
                    if size > 0:
                        self.stats["bytes"] += size
                    if res_type == "OK_IMG":
                        cm[path] = cell
                        self.log("[OK]", fu, f"Скачано {size/(1024*1024):.1f} МБ")
                    elif res_type == "OK_PDF":
                        cm[path] = cell
                        self.log("[OK]", fu, "PDF сохранён")
                    elif res_type == "OK_PDF_IMG":
                        for pp in path:
                            cm[pp] = cell
                        self.log("[OK]", fu, f"PDF → {len(path)} фото")
                    else:
                        status_code = "404" if err and ("404" in err or "Status 404" in err) else "ERROR"
                        # Если была отмена — не логируем как ошибку
                        if not self.check_cancel():
                            self.download_errors.append({
                                "url": fu, "status": status_code, "message": err, "folder": folder_name
                            })
                            try:
                                cell.fill = PatternFill(start_color="FF6347", fill_type="solid")
                            except Exception:
                                pass
                            self.log("[ОШИБКА]", fu, err)
                            self.stats["fail"] += 1
                except Exception as e:
                    if not self.check_cancel():
                        self.download_errors.append({
                            "url": val, "status": "ERROR", "message": str(e), "folder": folder_name
                        })
                        self.log("[ОШИБКА]", val, str(e))
                        self.stats["fail"] += 1
                proc += 1
                self._dl_done = proc

                # Progress: 5% (submit) + 65% (download) = максимум 70%
                # Показываем плавно даже если dl_submitted == 0
                if self._dl_submitted > 0:
                    dl_pct = 5 + min(65, int(proc * 65 / self._dl_submitted))
                else:
                    dl_pct = DL_PROGRESS_MAX
                if dl_pct != self._last_progress:
                    self.progress.emit(dl_pct)
                    self._last_progress = dl_pct

                # Heartbeat: каждые ~500 мс лог "в процессе: N из M"
                now = time.time()
                if (now - self._last_log_time) > 0.5:
                    self.log("[ИНФО]", "Загрузка",
                             f"Скачано {proc}/{self._dl_submitted}")
                    self._last_log_time = now
                    QApplication.processEvents()

                self.emit_stats()
                QApplication.processEvents()
        finally:
            # Закрываем executor в finally — чтобы cancel() или exception не оставил висящие потоки
            try:
                if not self._c:
                    self.executor.shutdown(wait=True)
                # Если была отмена — cancel() уже вызвал shutdown(wait=False, cancel_futures=True)
            except Exception:
                pass
            self.executor = None

        self.log("[ИНФО]", "Обработка", "Начинаю обработку...")
        thr = self.s.white_threshold
        ftp = [
            os.path.join(r, f)
            for r, _, fs in os.walk(dl_dir)
            for f in fs
            if is_processable_file(f)
        ]
        for i, rp in enumerate(ftp):
            if self.check_cancel():
                break
            self.check_pause()
            rlp = os.path.relpath(rp, dl_dir)
            fd = os.path.join(pr_dir, os.path.dirname(rlp))
            os.makedirs(fd, exist_ok=True)
            if rlp not in cm:
                continue
            cell = cm[rlp]
            if self.s.rename_mode == "article":
                fn = os.path.basename(os.path.dirname(rlp))
                sa = fn.split(" - ", 1)[-1] if " - " in fn else fn
                rn = os.path.splitext(os.path.basename(rlp))[0]
                ip = rn.split("_")[-1]
                ext = os.path.splitext(rlp)[1]
                fname = f"{sa}_{ip}{ext}"
            else:
                fname = os.path.basename(rlp)
            fp = os.path.join(fd, fname)
            if rp.lower().endswith(".pdf"):
                if not os.path.exists(fp):
                    shutil.copy2(rp, fp)
                cell.fill = PatternFill(start_color="90EE90", fill_type="solid")
                self.stats["ok"] += 1
                self.log("[OK]", fname, "PDF скопирован")
                self.emit_stats()
                self.progress.emit(70 + int((i + 1) / len(ftp) * 30))
                continue
            if os.path.exists(fp):
                cell.fill = PatternFill(start_color="90EE90", fill_type="solid")
                self.stats["ok"] += 1
                self.emit_stats()
                self.progress.emit(70 + int((i + 1) / len(ftp) * 30))
                continue
            try:
                with PILImage.open(rp) as img:
                    img = img.convert("RGBA")
                    pi, dp, cat = self.processor.process_with_analysis(img, self.s, threshold=thr)
                self.save_img(pi, os.path.join(fd, f"{dp}{fname}"))
                if cat == "SIZE_FAIL":
                    cell.fill = PatternFill(start_color="FFD700", fill_type="solid")
                    self.stats["defect_size"] += 1
                    self.log("[ОТБРАКОВАНО]", fname, "!РАЗМЕР_")
                elif cat == "QUESTION_WHITE":
                    cell.fill = PatternFill(start_color="87CEEB", fill_type="solid")
                    self.stats["defect_white"] += 1
                    self.log("[ПОД ВОПРОСОМ]", fname, "!БЕЛЫЙ_")
                else:
                    cell.fill = PatternFill(start_color="90EE90", fill_type="solid")
                    self.stats["ok"] += 1
                    self.log("[OK]", fname, "Успешно")
                del pi
            except Exception as e:
                cell.fill = PatternFill(start_color="FF6347", fill_type="solid")
                self.stats["fail"] += 1
                self.log("[ОШИБКА]", rp, str(e))
            self.emit_stats()
            self.progress.emit(70 + int((i + 1) / len(ftp) * 30))
        try:
            if self.s.report_copy:
                wb.save(os.path.join(os.path.dirname(self.s.source), f"Отчет_{os.path.basename(self.s.source)}"))
            else:
                wb.save(self.s.source)
            xl.close()
        except Exception as e:
            xl.close()
            raise Exception(f"Ошибка Excel: {e}")
        if self.s.clean_raw and os.path.exists(dl_dir):
            try:
                shutil.rmtree(dl_dir)
            except OSError:
                pass


# ============================================================
# VALERA PIXMAP CREATOR
# ============================================================
def create_valera_pixmap(size: int = Defaults.THUMBNAIL_SIZE):
    exe_dir = os.path.dirname(os.path.abspath(sys.argv[0]))
    valera_path = os.path.join(exe_dir, "valera.png")
    if os.path.exists(valera_path):
        pix = QPixmap(valera_path)
        if not pix.isNull():
            return pix.scaled(size, size, Qt.KeepAspectRatio, Qt.SmoothTransformation)
    pix = QPixmap(size, int(size * 1.1))
    pix.fill(Qt.transparent)
    with QPainter(pix) as p:
        p.setRenderHint(QPainter.Antialiasing)
        s = size / 50.0
        p.setBrush(QBrush(QColor(255, 193, 7)))
        p.setPen(QPen(QColor(230, 160, 0), max(1, int(s))))
        p.drawRoundedRect(int(10*s), int(1*s), int(30*s), int(12*s), int(4*s), int(4*s))
        p.drawRect(int(7*s), int(10*s), int(36*s), int(4*s))
        p.setBrush(QBrush(QColor(255, 224, 189)))
        p.setPen(QPen(QColor(210, 180, 150), max(1, int(s))))
        p.drawEllipse(int(14*s), int(12*s), int(22*s), int(20*s))
        p.setBrush(QBrush(QColor(50, 50, 50)))
        p.setPen(Qt.NoPen)
        p.drawEllipse(int(19*s), int(20*s), int(3*s), int(3*s))
        p.drawEllipse(int(28*s), int(20*s), int(3*s), int(3*s))
        p.setBrush(QBrush(QColor(255, 255, 255)))
        p.drawEllipse(int(20*s), int(20*s), max(1, int(1.5*s)), max(1, int(1.5*s)))
        p.drawEllipse(int(29*s), int(20*s), max(1, int(1.5*s)), max(1, int(1.5*s)))
        p.setPen(QPen(QColor(180, 100, 80), max(1, int(1.5*s))))
        p.setBrush(Qt.NoBrush)
        p.drawArc(int(20*s), int(25*s), int(10*s), int(6*s), 0, -180*16)
        p.setBrush(QBrush(QColor(76, 163, 224)))
        p.setPen(QPen(QColor(50, 130, 190), max(1, int(s))))
        p.drawRoundedRect(int(16*s), int(32*s), int(18*s), int(15*s), int(2*s), int(2*s))
        p.drawRect(int(8*s), int(34*s), int(8*s), int(5*s))
        p.drawRect(int(34*s), int(34*s), int(8*s), int(5*s))
        p.setBrush(QBrush(QColor(255, 224, 189)))
        p.setPen(QPen(QColor(210, 180, 150), max(1, int(s))))
        p.drawEllipse(int(6*s), int(33*s), int(6*s), int(6*s))
        p.drawEllipse(int(38*s), int(33*s), int(6*s), int(6*s))
        p.setBrush(QBrush(QColor(90, 90, 90)))
        p.setPen(QPen(QColor(60, 60, 60), max(1, int(s))))
        p.drawRect(int(17*s), int(46*s), int(7*s), int(7*s))
        p.drawRect(int(26*s), int(46*s), int(7*s), int(7*s))
        p.setBrush(QBrush(QColor(70, 50, 30)))
        p.setPen(Qt.NoPen)
        p.drawRoundedRect(int(15*s), int(51*s), int(9*s), int(4*s), int(2*s), int(2*s))
        p.drawRoundedRect(int(26*s), int(51*s), int(9*s), int(4*s), int(2*s), int(2*s))
    return pix


# ============================================================
# PREVIEW WIDGETS
# ============================================================
class ThumbnailWidget(QFrame):
    selection_changed = Signal()

    # Размер увеличен на 30% (был 110x140 / 90x90)
    THUMB_W = 143
    THUMB_H = 160
    IMG_W = 117
    IMG_H = 117

    def __init__(self, filepath: str, category: str, parent=None):
        super().__init__(parent)
        self.filepath = filepath
        self.category = category
        self.setFixedSize(self.THUMB_W, self.THUMB_H)
        self.setFrameStyle(QFrame.StyledPanel)
        self._update_style(False)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(4, 4, 4, 4)
        layout.setSpacing(2)

        # ── Картинка ──
        self.img_label = QLabel()
        self.img_label.setFixedSize(self.IMG_W, self.IMG_H)
        self.img_label.setAlignment(Qt.AlignCenter)
        self.img_label.setStyleSheet("background: #eee; border: 1px solid #ccc;")
        self._load_thumbnail()
        layout.addWidget(self.img_label, alignment=Qt.AlignCenter)

        # ── Нижний ряд: [✓] + имя файла ──
        bottom = QHBoxLayout()
        bottom.setSpacing(2)

        self.checkbox = QCheckBox()
        self.checkbox.setStyleSheet("QCheckBox::indicator { width: 22px; height: 22px; }")
        self.checkbox.stateChanged.connect(lambda: self._update_style(self.checkbox.isChecked()))
        self.checkbox.stateChanged.connect(self.selection_changed)
        bottom.addWidget(self.checkbox)

        name = os.path.basename(filepath)
        dn = name[:20] + "…" if len(name) > 20 else name
        self.name_label = QLabel(dn)
        self.name_label.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        self.name_label.setStyleSheet("font-size: 12px;")   # ← было 10 px, +20 %
        self.name_label.setWordWrap(True)
        self.name_label.setFixedHeight(28)
        bottom.addWidget(self.name_label, 1)

        layout.addLayout(bottom)

        # Двойной клик по превью → открыть файл
        self.img_label.mouseDoubleClickEvent = lambda e: QDesktopServices.openUrl(QUrl.fromLocalFile(filepath))
        self._setup_tooltip()

    def _setup_tooltip(self):
        try:
            with PILImage.open(self.filepath) as im:
                w, h = im.size
            ext = os.path.splitext(self.filepath)[1].upper().lstrip(".")
            self.setToolTip(f"{w} x {h} px\n{ext}")
        except Exception:
            self.setToolTip("Файл")

    def _load_thumbnail(self):
        try:
            if self.filepath.lower().endswith(".pdf"):
                self.img_label.setText("📄 PDF")
                self.img_label.setStyleSheet(
                    "background: #e0f0fb; border: 1px solid #4CA3E0; font-weight: bold; font-size: 20px; color: #2C3E50;"
                )
                return
            pm = QPixmap(self.filepath)
            if not pm.isNull():
                self.img_label.setPixmap(pm.scaled(self.IMG_W, self.IMG_H, Qt.KeepAspectRatio, Qt.SmoothTransformation))
            else:
                self.img_label.setText("Файл")
        except Exception:
            self.img_label.setText("Ошибка")

    def _update_style(self, sel):
        if sel:
            self.setStyleSheet("QFrame { border: 3px solid #4CA3E0; border-radius: 4px; background: #e0f0fb; }")
        else:
            self.setStyleSheet("QFrame { border: 1px solid #ccc; border-radius: 4px; background: white; }")

    def is_selected(self):
        return self.checkbox.isChecked()

    def set_selected(self, v):
        self.checkbox.setChecked(v)


class DownloadErrorThumbnail(QFrame):
    # Размер увеличен на 30% (был 110x140 / 90x90)
    THUMB_W = 143
    THUMB_H = 182
    IMG_W = 117
    IMG_H = 117

    def __init__(self, url, status, message, parent=None):
        super().__init__(parent)
        url_str = str(url or "Неизвестный URL")
        status_str = str(status or "ОШИБКА")
        message_str = str(message or "")
        self.url = url_str
        self.setFixedSize(self.THUMB_W, self.THUMB_H)
        self.setFrameStyle(QFrame.StyledPanel)
        self.setStyleSheet("QFrame { border: 2px solid #E74C3C; border-radius: 4px; background: #FDEDEC; }")
        self.setCursor(Qt.PointingHandCursor)
        layout = QVBoxLayout(self)
        layout.setContentsMargins(4, 4, 4, 4)
        layout.setSpacing(2)

        self.img_label = QLabel()
        self.img_label.setFixedSize(self.IMG_W, self.IMG_H)
        self.img_label.setAlignment(Qt.AlignCenter)
        self.img_label.setStyleSheet("background: #FADBD8; border: 1px solid #E74C3C; border-radius: 3px;")
        pix = QPixmap(75, 75)
        pix.fill(Qt.transparent)
        with QPainter(pix) as p:
            p.setRenderHint(QPainter.Antialiasing)
            p.setPen(QPen(QColor(231, 76, 60), 7))
            p.drawLine(15, 15, 60, 60)
            p.drawLine(60, 15, 15, 60)
        self.img_label.setPixmap(pix)
        layout.addWidget(self.img_label, alignment=Qt.AlignCenter)

        self.status_label = QLabel(f"❌ {status_str}")
        self.status_label.setAlignment(Qt.AlignCenter)
        self.status_label.setStyleSheet("color: #C0392B; font-size: 10px; font-weight: bold;")
        self.status_label.setWordWrap(True)
        self.status_label.setFixedHeight(16)
        layout.addWidget(self.status_label)

        display = url_str if len(url_str) <= 32 else url_str[:32] + "…"
        self.url_label = QLabel(display)
        self.url_label.setAlignment(Qt.AlignCenter)
        self.url_label.setStyleSheet("color: #7F8C8D; font-size: 9px;")
        self.url_label.setWordWrap(True)
        self.url_label.setFixedHeight(24)
        layout.addWidget(self.url_label)

        self.img_label.mouseDoubleClickEvent = lambda e: QDesktopServices.openUrl(QUrl.fromUserInput(url_str))
        self.mouseDoubleClickEvent = lambda e: QDesktopServices.openUrl(QUrl.fromUserInput(url_str))
        safe_msg = escape(message_str)
        self.setToolTip(f"❌ Ошибка загрузки\nСтатус: {status_str}\n{safe_msg}\n\nДвойной клик — открыть в браузере")


class ProductRowWidget(QFrame):
    def __init__(self, fn, parent=None):
        super().__init__(parent)
        self.folder_name = fn
        self.thumbs = []
        self.setFrameStyle(QFrame.StyledPanel)
        self.setStyleSheet("QFrame { border: 1px solid #ddd; border-radius: 4px; background: #fafafa; padding: 4px; }")
        layout = QVBoxLayout(self)
        layout.setContentsMargins(5, 5, 5, 5)
        layout.setSpacing(4)
        self.title_lbl = QLabel(f"📦 {fn}")
        layout.addWidget(self.title_lbl)
        self.tl = QHBoxLayout()
        self.tl.setSpacing(10)
        self.tl.setAlignment(Qt.AlignLeft)
        layout.addLayout(self.tl)

    def add_widget(self, t):
        self.thumbs.append(t)
        self.tl.addWidget(t)


class ProductPreviewDialog(QDialog):
    PAGE_SIZE = Defaults.PREVIEW_PAGE_SIZE

    def __init__(self, folder_path, settings, parent=None, download_errors=None, threshold_level=0):
        super().__init__(parent)
        self.setWindowTitle("Предпросмотр результатов")
        self.resize(1100, 750)
        self.setWindowFlags(self.windowFlags() | Qt.WindowMaximizeButtonHint)
        self.folder_path = folder_path
        self.s = settings
        self.product_rows = []
        self.all_data = []
        self.filtered_data = []
        self.current_page = 0
        self.download_errors = download_errors or []
        self.current_threshold = max(80, Defaults.THRESHOLD_WHITE - (threshold_level * 20))
        self.undo_mgr = UndoManager(max_steps=Defaults.MAX_UNDO_STEPS)
        undo_dir = os.path.join(os.path.dirname(folder_path), ".valera_undo_" + os.path.basename(folder_path))
        self.undo_mgr.init_temp(undo_dir)
        self.shortcut_undo = QShortcut(QKeySequence("Ctrl+Z"), self)
        self.shortcut_undo.activated.connect(self.do_undo)

        self._build_ui()
        self._load_records()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(8, 8, 8, 8)   # ← чуть ужали
        layout.setSpacing(3)                    # ← было ~6, отступы вдвое меньше

        # УБРАНО: верхний лейбл «📦 Галочки ➔ Действие снизу...»

        self.sa = QScrollArea()
        self.sa.setWidgetResizable(True)
        self.lw = QWidget()
        self.ll = QVBoxLayout(self.lw)
        self.ll.setSpacing(10)
        self.ll.setAlignment(Qt.AlignTop)
        self.sa.setWidget(self.lw)
        layout.addWidget(self.sa)

        self.cl = QLabel("Выбрано: 0")
        self.cl.setStyleSheet("font-weight: bold; font-size: 12px;")
        layout.addWidget(self.cl)

        # ── Пагинация ──
        pl = QHBoxLayout()
        pl.setSpacing(4)
        self.btn_prev = QPushButton("◀ Назад")
        self.btn_prev.clicked.connect(self._prev_page)
        self.lbl_page = QLabel("1 / 1")
        self.lbl_page.setAlignment(Qt.AlignCenter)
        self.lbl_page.setMinimumWidth(80)
        self.lbl_page.setStyleSheet("font-weight: bold;")
        self.btn_next = QPushButton("Вперед ▶")
        self.btn_next.clicked.connect(self._next_page)
        pl.addStretch()
        pl.addWidget(self.btn_prev)
        pl.addWidget(self.lbl_page)
        pl.addWidget(self.btn_next)
        pl.addStretch()
        layout.addLayout(pl)

        # ── Прогресс операции ──
        self.op_progress_bar = QProgressBar()
        self.op_progress_bar.setFixedHeight(14)
        self.op_progress_bar.setMaximum(100)
        self.op_progress_bar.setTextVisible(False)
        self.op_progress_bar.setStyleSheet(
            "QProgressBar { border: 1px solid #ccc; border-radius: 3px; background: #eee; }"
            "QProgressBar::chunk { background: #4CA3E0; border-radius: 2px; }"
        )
        self.op_progress_bar.setVisible(False)
        self.op_progress_label = QLabel("")
        self.op_progress_label.setStyleSheet("color: #4CA3E0; font-size: 11px;")
        op_progress_layout = QHBoxLayout()
        op_progress_layout.addWidget(self.op_progress_bar)
        op_progress_layout.addWidget(self.op_progress_label)
        layout.addLayout(op_progress_layout)

        # ── Нижний ряд кнопок ──
        bl = QHBoxLayout()
        bl.setSpacing(6)
        btn_min_h = 32      # ← было 40, −20 %
        btn_min_w = 120     # ← было 150, −20 %

        b_undo = QPushButton("↩ Отмена (Ctrl+Z)")
        b_undo.setMinimumSize(btn_min_w, btn_min_h)
        b_undo.clicked.connect(self.do_undo)

        b_crop = QPushButton("✂ Обрезать поля")
        b_crop.setMinimumSize(btn_min_w, btn_min_h)
        b_crop.clicked.connect(self.action_crop)

        b_crop_sq = QPushButton("⬜ Обрезать в квадрат")
        b_crop_sq.setMinimumSize(btn_min_w, btn_min_h)
        b_crop_sq.clicked.connect(self.action_crop_square)

        b_sq = QPushButton("⬜ Центрировать")
        b_sq.setMinimumSize(btn_min_w, btn_min_h)
        b_sq.clicked.connect(self.action_square)

        b_reduce = QPushButton("🔍 Сократить отступ")
        b_reduce.setMinimumSize(btn_min_w, btn_min_h)
        b_reduce.clicked.connect(self.action_reduce_padding)

        b_orig = QPushButton("↩ Исходник")
        b_orig.setMinimumSize(btn_min_w, btn_min_h)
        b_orig.clicked.connect(self.action_original)

        b_del = QPushButton("🗑 Удалить")
        b_del.setMinimumSize(btn_min_w, btn_min_h)
        b_del.setStyleSheet("QPushButton{color:red;font-weight:bold}")
        b_del.clicked.connect(self.action_delete)

        b_cancel = QPushButton("Закрыть")
        b_cancel.setMinimumSize(100, btn_min_h)
        b_cancel.clicked.connect(self.reject)

        bl.addWidget(b_undo)
        bl.addWidget(b_crop)
        bl.addWidget(b_crop_sq)
        bl.addWidget(b_sq)
        bl.addWidget(b_reduce)
        bl.addWidget(b_orig)
        bl.addWidget(b_del)
        bl.addStretch()
        bl.addWidget(b_cancel)
        layout.addLayout(bl)

    def do_undo(self):
        ok, msg = self.undo_mgr.undo()
        if ok:
            self._load_records()
            self._update_count()
            QMessageBox.information(self, "Отмена", msg)
        else:
            QMessageBox.warning(self, "Отмена", msg)

    def _render_page(self):
        while self.ll.count():
            item = self.ll.takeAt(0)
            w = item.widget()
            if w:
                w.deleteLater()
        self.product_rows.clear()
        total_pages = max(1, (len(self.filtered_data) + self.PAGE_SIZE - 1) // self.PAGE_SIZE)
        self.current_page = min(self.current_page, total_pages - 1)
        start = self.current_page * self.PAGE_SIZE
        end = min(start + self.PAGE_SIZE, len(self.filtered_data))
        self.lw.setUpdatesEnabled(False)
        for item in self.filtered_data[start:end]:
            rw = ProductRowWidget(item["name"], self)
            for fp, cat in item.get("files", []):
                t = ThumbnailWidget(fp, cat, self)
                t.selection_changed.connect(self._update_count)
                rw.add_widget(t)
            for err in item.get("errors", []):
                et = DownloadErrorThumbnail(err.get("url"), err.get("status"), err.get("message"), self)
                rw.add_widget(et)
            self.product_rows.append(rw)
            self.ll.addWidget(rw)
        self.lw.setUpdatesEnabled(True)
        self._update_pagination_controls(total_pages)

    def _update_pagination_controls(self, total_pages: int):
        self.lbl_page.setText(f"{self.current_page + 1} / {total_pages}")
        self.btn_prev.setEnabled(self.current_page > 0)
        self.btn_next.setEnabled(self.current_page < total_pages - 1)

    def _prev_page(self):
        if self.current_page > 0:
            self.current_page -= 1
            self._render_page()
            self._update_count()

    def _next_page(self):
        total_pages = max(1, (len(self.filtered_data) + self.PAGE_SIZE - 1) // self.PAGE_SIZE)
        if self.current_page < total_pages - 1:
            self.current_page += 1
            self._render_page()
            self._update_count()

    def _get_selected(self):
        return [
            t.filepath
            for row in self.product_rows
            for t in row.thumbs
            if isinstance(t, ThumbnailWidget) and t.is_selected() and t.isVisible()
        ]
    def _process_edit_action(self, files, process_func):
        if not files:
            return
        self.op_progress_bar.setVisible(True)
        self.op_progress_bar.setValue(0)
        self.op_progress_label.setText("⏳ Выполняется...")
        QApplication.processEvents()
        proc = ImageProcessor()
        ops = []
        for fp in files:
            backup = self.undo_mgr.backup_for_edit(fp)
            if backup:
                ops.append(("edit", fp, backup))
        if not ops:
            self.op_progress_bar.setVisible(False)
            self.op_progress_label.setText("")
            QMessageBox.warning(self, "Отмена", "Не удалось создать резервные копии")
            return
        for i, fp in enumerate(files):
            self.op_progress_bar.setValue(int(((i + 1) / len(files)) * 100))
            QApplication.processEvents()
            try:
                if fp.lower().endswith(".pdf"):
                    continue
                with PILImage.open(fp) as img:
                    img = img.convert("RGBA")
                    img_final = process_func(proc, img)
                if self.s.replace_transparent and img_final.mode == "RGBA":
                    bg = PILImage.new("RGB", img_final.size, (255, 255, 255))
                    bg.paste(img_final, mask=img_final.split()[3])
                    img_final = bg
                img_final.save(fp)
            except Exception as e:
                LOGGER.error("Editor error: %s", e)
        self.undo_mgr.push_batch(ops)
        self.current_threshold = max(80, self.current_threshold - 20)
        self.op_progress_bar.setVisible(False)
        self.op_progress_label.setText("")
        self._load_records()
        gc.collect()

    def action_crop(self):
        def process_func(proc, img):
            c = proc._crop_to_content(img, self.current_threshold)
            if c:
                img = c
            img_final, _ = proc.process(img, self.s, "white_bg", skip_center=True)
            return img_final
        self._process_edit_action(self._get_selected(), process_func)

    def action_crop_square(self):
        def process_func(proc, img):
            sq = proc._smart_crop_to_square(img, self.current_threshold)
            img_final, _ = proc.process(sq, self.s, "white_bg", skip_center=True)
            return img_final
        self._process_edit_action(self._get_selected(), process_func)

    def action_square(self):
        def process_func(proc, img):
            c = proc._crop_to_content(img, self.current_threshold)
            if c:
                img = c
            sq = proc._center_in_square(img, self.s.padding_pct)
            img_final, _ = proc.process(sq, self.s, "white_bg", skip_center=True)
            return img_final
        self._process_edit_action(self._get_selected(), process_func)

    def action_reduce_padding(self):
        def process_func(proc, img):
            c = proc._crop_to_content(img, self.current_threshold)
            if c:
                img = c
            sq = proc._center_in_square(img, max(Defaults.MIN_PADDING, self.s.padding_pct // 3))
            img_final, _ = proc.process(sq, self.s, "white_bg", skip_center=True)
            return img_final
        self._process_edit_action(self._get_selected(), process_func)

    def action_original(self):
        files = self._get_selected()
        if not files:
            return
        raw_dir = self.folder_path.replace("обработано_", "скачано_")
        if not os.path.exists(raw_dir):
            QMessageBox.warning(self, "Ошибка", "Папка с исходниками не найдена!")
            return
        self.op_progress_bar.setVisible(True)
        self.op_progress_bar.setValue(0)
        self.op_progress_label.setText("⏳ Восстановление...")
        QApplication.processEvents()
        ops = []
        for fp in files:
            backup = self.undo_mgr.backup_for_edit(fp)
            if backup:
                ops.append(("edit", fp, backup))
        if not ops:
            self.op_progress_bar.setVisible(False)
            self.op_progress_label.setText("")
            return
        for i, fp in enumerate(files):
            self.op_progress_bar.setValue(int(((i + 1) / len(files)) * 100))
            QApplication.processEvents()
            rel = os.path.relpath(fp, self.folder_path)
            raw_fp = os.path.join(raw_dir, rel)
            raw_folder = os.path.dirname(raw_fp)
            if os.path.exists(raw_folder):
                raws = [f for f in os.listdir(raw_folder) if f.startswith("raw_")]
                if raws:
                    try:
                        shutil.copy2(os.path.join(raw_folder, raws[0]), fp)
                    except Exception as e:
                        LOGGER.error("Restore original error: %s", e)
        self.undo_mgr.push_batch(ops)
        self.op_progress_bar.setVisible(False)
        self.op_progress_label.setText("")
        self._load_records()

    def action_delete(self):
        files = self._get_selected()
        if not files:
            return
        if QMessageBox.question(self, "Удаление", f"Удалить {len(files)} файлов? (можно отменить)") == QMessageBox.Yes:
            ops = []
            for fp in files:
                backup = self.undo_mgr.backup_for_delete(fp)
                if backup:
                    ops.append(("delete", fp, backup))
            if ops:
                self.undo_mgr.push_batch(ops)
            self._load_records()

    def _load_records(self):
        self.all_data.clear()
        st = defaultdict(list)
        for root, _, fns in os.walk(self.folder_path):
            for fn in fns:
                if not is_processable_file(fn):
                    continue
                fp = os.path.join(root, fn)
                rel = os.path.relpath(fp, self.folder_path)
                parts = rel.split(os.sep)
                folder_name = parts[1] if len(parts) >= 3 else (parts[0] if len(parts) == 2 else "Без группы")
                cat = "OK_INTERIOR"
                if fn.startswith("!БЕЛЫЙ_"):
                    cat = "QUESTION_WHITE"
                elif fn.startswith("!РАЗМЕР_"):
                    cat = "SIZE_FAIL"
                elif fn.startswith("!ОШИБКА_"):
                    cat = "ERROR"
                st[folder_name].append((fp, cat))
        err_by_folder = defaultdict(list)
        for err in self.download_errors:
            folder = err.get("folder", "Без группы")
            err_by_folder[folder].append(err)
        all_folders = set(st.keys()) | set(err_by_folder.keys())
        for fn in sorted(all_folders):
            files = st.get(fn, [])
            errors = err_by_folder.get(fn, [])
            cats = set(cat for _, cat in files)
            if errors:
                cats.add("DOWNLOAD_ERROR")
            self.all_data.append({
                "type": "product", "name": fn, "files": files,
                "errors": errors, "categories": cats
            })
        self.filtered_data = self.all_data[:]
        self.current_page = 0
        self._render_page()
        self._update_count()

    def _update_count(self):
        total = len(self._get_selected())
        self.cl.setText(f"Выбрано: {total}")

    def select_all(self):
        for r in self.product_rows:
            if r.isVisible():
                for t in r.thumbs:
                    if isinstance(t, ThumbnailWidget):
                        t.set_selected(True)

    def deselect_all(self):
        for r in self.product_rows:
            for t in r.thumbs:
                if isinstance(t, ThumbnailWidget):
                    t.set_selected(False)


# ============================================================
# LOG DIALOG & PROGRESS WIDGET
# ============================================================
class LogDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Лог")
        self.setMinimumSize(1000, 700)
        lay = QVBoxLayout(self)
        self.table = QTableWidget()
        self.table.setColumnCount(3)
        self.table.setHorizontalHeaderLabels(["Статус", "Источник", "Сообщение"])
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Interactive)
        self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.setColumnWidth(1, 500)
        lay.addWidget(self.table)
        b = QPushButton("Закрыть")
        b.clicked.connect(self.close)
        lay.addWidget(b)

    def add_rows_batch(self, rows):
        self.table.setUpdatesEnabled(False)
        for s, so, m in rows:
            r = self.table.rowCount()
            self.table.insertRow(r)
            for c, t in enumerate([s, so, m]):
                self.table.setItem(r, c, QTableWidgetItem(t))
        self.table.setUpdatesEnabled(True)
        self.table.scrollToBottom()


class ValeraProgressWidget(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setMinimumHeight(70)
        self._v = 0
        self._valera_pix = create_valera_pixmap(Defaults.PROGRESS_VALERA_SIZE)

    def setValue(self, v: int):
        self._v = max(0, min(100, v))
        self.update()

    def paintEvent(self, e):
        with QPainter(self) as p:
            p.setRenderHint(QPainter.Antialiasing)
            track_y = 28
            track_h = 18
            margin = 55
            track_rect = QRect(margin, track_y, self.width() - margin * 2, track_h)

            p.setBrush(QBrush(QColor(220, 220, 220)))
            p.setPen(QPen(QColor(200, 200, 200), 1))
            p.drawRoundedRect(track_rect, 9, 9)

            if self._v > 0:
                fill_w = int(track_rect.width() * (self._v / 100.0))
                fill_rect = QRect(track_rect.x(), track_rect.y(), max(fill_w, track_h), track_h)
                grad = QLinearGradient(fill_rect.topLeft(), fill_rect.topRight())
                grad.setColorAt(0, QColor(76, 163, 224))
                grad.setColorAt(1, QColor(130, 210, 255))
                p.setBrush(QBrush(grad))
                p.setPen(Qt.NoPen)
                p.drawRoundedRect(fill_rect, 9, 9)

            vx = track_rect.x() + int(track_rect.width() * (self._v / 100.0)) - 24
            vx = max(track_rect.x() - 24, min(vx, track_rect.right() - 24))
            if self._valera_pix:
                p.drawPixmap(int(vx), track_y - 28, self._valera_pix)

            p.setPen(QPen(QColor(80, 80, 80)))
            f = p.font()
            f.setBold(True)
            f.setPointSize(10)
            p.setFont(f)
            p.drawText(
                QRect(0, track_y + track_h + 3, self.width(), 18),
                Qt.AlignCenter,
                f"Валера тащит: {self._v}%"
            )


# ============================================================
# AUTHOR DIALOG & MAIN WINDOW
# ============================================================
class AuthorDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Автор")
        self.setFixedSize(380, 220)
        self.setStyleSheet("QDialog { background: white; }")
        layout = QVBoxLayout(self)
        layout.setSpacing(12)
        layout.setContentsMargins(20, 20, 20, 15)
        vp = create_valera_pixmap(40)
        if not vp.isNull():
            lbl_img = QLabel()
            lbl_img.setPixmap(vp)
            lbl_img.setAlignment(Qt.AlignCenter)
            layout.addWidget(lbl_img)
        t1 = QLabel('Гитхаб автора: <a href="https://github.com/allkirill" style="color:#4CA3E0;">https://github.com/allkirill</a>')
        t1.setOpenExternalLinks(True)
        t1.setStyleSheet("font-size: 12px;")
        layout.addWidget(t1)
        t2 = QLabel('Еще несколько решений: <a href="https://vlookup-app.ru" style="color:#4CA3E0;">https://vlookup-app.ru</a>')
        t2.setOpenExternalLinks(True)
        t2.setStyleSheet("font-size: 12px;")
        layout.addWidget(t2)
        t3 = QLabel("Фотоукладчик Валера всегда на связи 🛠")
        t3.setStyleSheet("font-weight: bold; font-size: 13px; color: #4CA3E0; margin-top: 6px;")
        layout.addWidget(t3)
        layout.addStretch()
        btn = QPushButton("Закрыть")
        btn.setFixedWidth(100)
        btn.clicked.connect(self.close)
        layout.addWidget(btn, alignment=Qt.AlignCenter)


class MainWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Фотоукладчик Валера")
        self.setMinimumSize(760, 720)
        self.resize(860, 780)
        self.setWindowIcon(QIcon(create_valera_pixmap(32)))
        self.source_path = None
        self.output_dir = None
        self.settings_io = QSettings("ValeraSoft", "PhotoLoader")
        self.worker = None
        self.log_dialog = None
        self.setAcceptDrops(True)
        self._ap = False
        self._srf = []
        self.log_buffer = []
        self.log_timer = QTimer()
        self.log_timer.timeout.connect(self.flush_log)
        self.log_timer.start(300)
        self.last_download_errors = []
        self.last_processed_dir = None
        self.force_threshold_level = 0
        self.watermark_path = ""
        self.build_ui()
        self.load_settings()

    def build_ui(self):
        ml = QVBoxLayout(self)
        ml.setContentsMargins(10, 10, 10, 10)
        ml.setSpacing(6)
        self._build_source_section(ml)
        self._build_preset_section(ml)
        self._build_params_section(ml)
        self._build_white_bg_section(ml)
        self._build_log_section(ml)
        self._build_progress_section(ml)
        self._build_buttons_section(ml)
        self._build_author_section(ml)

    def _build_source_section(self, ml):
        tg = QGridLayout()
        tg.setSpacing(5)
        self.bi = QToolButton()
        self.bi.setText("ℹ")
        self.bi.setStyleSheet("border:none;color:blue;font-weight:bold;font-size:16px;")
        self.bi.clicked.connect(self.show_info)
        self.stc = QComboBox()
        self.stc.addItems(["Excel", "Папка"])
        self.stc.setFixedWidth(90)
        self.stc.activated.connect(self.on_stc)
        self.bs = QPushButton("Выбрать источник")
        self.bs.setFixedWidth(140)
        self.ls = QLabel("Не выбран")
        self.ls.setStyleSheet("color:gray;")
        self.ls.setWordWrap(True)
        tg.addWidget(self.bi, 0, 0)
        tg.addWidget(self.stc, 0, 1)
        tg.addWidget(self.bs, 0, 2)
        tg.addWidget(self.ls, 0, 3)

        cl = generate_columns()
        he = QHBoxLayout()
        he.setSpacing(4)
        self.ca = QComboBox()
        self.ca.addItems(cl)
        self.ca.setFixedWidth(50)
        self.ca.setEditable(True)
        self.cuf = QComboBox()
        self.cuf.addItems(cl)
        self.cuf.setFixedWidth(50)
        self.cuf.setEditable(True)
        self.cut = QComboBox()
        self.cut.addItems(cl)
        self.cut.setCurrentText("P")
        self.cut.setFixedWidth(50)
        self.cut.setEditable(True)
        self.cr = QComboBox()
        self.cr.addItems(["По артикулу", "Оригинальное"])
        self.cr.setFixedWidth(110)
        self.cfs = QComboBox()
        self.cfs.addItems(["В порядке Excel", "По алфавиту"])
        self.cfs.setFixedWidth(120)
        he.addWidget(QLabel("Артикул:"))
        he.addWidget(self.ca)
        he.addSpacing(6)
        he.addWidget(QLabel("Ссылки от:"))
        he.addWidget(self.cuf)
        he.addWidget(QLabel("до:"))
        he.addWidget(self.cut)
        he.addSpacing(6)
        he.addWidget(QLabel("Имя:"))
        he.addWidget(self.cr)
        he.addSpacing(6)
        he.addWidget(QLabel("Папки:"))
        he.addWidget(self.cfs)
        he.addStretch()
        self.weo = QWidget()
        self.weo.setLayout(he)
        tg.addWidget(self.weo, 1, 0, 1, 4)

        hr = QHBoxLayout()
        self.chk_rc = QCheckBox("Отчёт в копию")
        self.chk_cr = QCheckBox("Удалить исходники")
        self.chk_ssl = QCheckBox("SSL")
        self.chk_ssl.setChecked(True)
        self.chk_agg = QCheckBox("Искать фото в папках и HTML-страницах")
        self.chk_agg.setChecked(True)
        b1 = QToolButton()
        b1.setText("ℹ")
        b1.setStyleSheet("border:none;color:blue;")
        b1.clicked.connect(lambda: QMessageBox.information(self, "Справка", "Отчёт сохранится в копию, чтобы не сломать оригинал."))
        b2 = QToolButton()
        b2.setText("ℹ")
        b2.setStyleSheet("border:none;color:blue;")
        b2.clicked.connect(lambda: QMessageBox.information(self, "Справка", "Удалит папку 'скачано' после обработки."))
        b3 = QToolButton()
        b3.setText("ℹ")
        b3.setStyleSheet("border:none;color:blue;")
        b3.clicked.connect(lambda: QMessageBox.information(self, "Справка", "Отключите, если ошибки SSL в корпоративной сети."))
        b4 = QToolButton()
        b4.setText("ℹ")
        b4.setStyleSheet("border:none;color:blue;")
        b4.clicked.connect(lambda: QMessageBox.information(
            self, "Умный парсинг ссылок",
            "Когда включено:\n"
            "• Если ссылка ведёт на ПАПКУ (Google Drive, Mail.ru, Dropbox) — "
            "программа откроет её и скачает все найденные файлы.\n"
            "• Если ссылка ведёт на страницу-обёртку с превью (HTML вместо картинки) — "
            "попытается извлечь прямую ссылку на файл через og:image, meta-refresh, "
            "JSON-LD и другие подсказки.\n\n"
            "Когда выключено:\n"
            "• Ссылки на папки вызовут ошибку — программа скачает только то, что "
            "является файлом по прямой ссылке. Ничего лишнего не подтянется.\n"
            "• Для случаев, когда вы уверены, что все ссылки — прямые ссылки на фотографии, "
            "и не хотите рисковать."
        ))
        hr.addWidget(self.chk_rc)
        hr.addWidget(b1)
        hr.addSpacing(10)
        hr.addWidget(self.chk_cr)
        hr.addWidget(b2)
        hr.addSpacing(10)
        hr.addWidget(self.chk_ssl)
        hr.addWidget(b3)
        hr.addSpacing(10)
        hr.addWidget(self.chk_agg)
        hr.addWidget(b4)
        hr.addStretch()
        self.wer = QWidget()
        self.wer.setLayout(hr)
        tg.addWidget(self.wer, 2, 0, 1, 4)

        self.bd = QPushButton("Место назначения")
        self.bd.setFixedWidth(140)
        self.ld = QLabel("Рабочий стол")
        self.ld.setStyleSheet("color:gray;")
        tg.addWidget(QLabel(""), 3, 0)
        tg.addWidget(self.bd, 3, 2)
        tg.addWidget(self.ld, 3, 3)
        ml.addLayout(tg)

    def _build_preset_section(self, ml):
        pl = QHBoxLayout()
        pl.addWidget(QLabel("💡 Пресет:"))
        self.pc = QComboBox()
        self.pc.addItems(list(PRESETS.keys()))
        self.pc.activated.connect(self.on_ps)
        pl.addWidget(self.pc)
        pl.addStretch()
        ml.addLayout(pl)

    def _build_params_section(self, ml):
        gp = QGroupBox("Параметры")
        gl = QHBoxLayout(gp)
        gl.setSpacing(5)
        self.fb = QComboBox()
        self.fb.addItems(["jpg", "png", "webp"])
        self.fb.setFixedWidth(60)
        self.al = QComboBox()
        self.al.addItems(["По высоте", "По ширине"])
        self.al.setFixedWidth(85)
        self.mi = QLineEdit("0")
        self.mi.setFixedWidth(45)
        self.mi.setValidator(QIntValidator(0, Defaults.MAX_PX_CAP, self))
        self.ma = QLineEdit("4000")
        self.ma.setFixedWidth(45)
        self.ma.setValidator(QIntValidator(0, Defaults.MAX_PX_CAP, self))
        self.up = QLineEdit("50")
        self.up.setFixedWidth(35)
        self.up.setValidator(QIntValidator(0, Defaults.MAX_UPSCALE_CAP, self))
        bu = QToolButton()
        bu.setText("ℹ")
        bu.setStyleSheet("border:none;color:blue;")
        bu.clicked.connect(lambda: QMessageBox.information(self, "Справка", "Если % увеличения больше лимита, фото пометится !РАЗМЕР_"))
        gl.addWidget(QLabel("Формат:"))
        gl.addWidget(self.fb)
        gl.addSpacing(4)
        gl.addWidget(self.al)
        gl.addSpacing(4)
        gl.addWidget(QLabel("Мин px:"))
        gl.addWidget(self.mi)
        gl.addSpacing(4)
        gl.addWidget(QLabel("Макс px:"))
        gl.addWidget(self.ma)
        gl.addSpacing(4)
        gl.addWidget(QLabel("Макс.увел %:"))
        gl.addWidget(self.up)
        gl.addWidget(bu)
        gl.addStretch()
        ml.addWidget(gp)

        # Вторая строка параметров: водный знак + очистка метаданных
        wm_row = QHBoxLayout()
        self.btn_wm = QPushButton("Наложить водный знак")
        self.btn_wm.setFixedWidth(160)
        self.btn_wm.setCursor(Qt.PointingHandCursor)
        self.btn_wm.clicked.connect(self.pick_watermark)
        self.btn_wm_clear = QPushButton("✕")
        self.btn_wm_clear.setFixedWidth(22)
        self.btn_wm_clear.setVisible(False)
        self.btn_wm_clear.setStyleSheet(
            "QPushButton{color:red;font-weight:bold;border:1px solid #ccc;border-radius:3px}"
            "QPushButton:hover{background:#ffcccc}"
        )
        self.btn_wm_clear.clicked.connect(self.clear_watermark)
        self.lbl_wm = QLabel("")
        self.lbl_wm.setStyleSheet("color:gray;font-size:10px;")
        
        self.chk_rm = QCheckBox("Очистить метаданные")
        
        wm_row.addWidget(self.btn_wm)
        wm_row.addWidget(self.btn_wm_clear)
        wm_row.addWidget(self.lbl_wm)
        wm_row.addStretch()
        wm_row.addWidget(self.chk_rm)
        wm_row.addStretch()
        ml.addLayout(wm_row)

    def _build_white_bg_section(self, ml):
        gw = QGroupBox("Белый фон и PDF")
        wl = QVBoxLayout(gw)
        
        # Главная галочка
        top_row = QHBoxLayout()
        self.chk_process_white = QCheckBox("Обрабатывать фото на белом фоне")
        self.chk_process_white.setChecked(True)
        self.chk_process_white.clicked.connect(self._on_white_bg_toggled)
        top_row.addWidget(self.chk_process_white)
        top_row.addStretch()
        wl.addLayout(top_row)
        
        # Подчинённые настройки
        self.white_bg_container = QWidget()
        sub_layout = QHBoxLayout(self.white_bg_container)
        sub_layout.setContentsMargins(20, 0, 0, 0)
        sub_layout.setSpacing(5)
        
        self.chk_cs = QCheckBox("Центрировать в квадрат")
        self.pi = QLineEdit("10")
        self.pi.setFixedWidth(35)
        self.pi.setValidator(QIntValidator(0, 100, self))
        bpi = QToolButton()
        bpi.setText("ℹ")
        bpi.setStyleSheet("border:none;color:blue;")
        bpi.clicked.connect(lambda: QMessageBox.information(
            self, "Справка",
            "Процент отступа вокруг объекта при центрировании в квадрат.\n"
            "При повторном нажатии 'Принудительно преобразовать' порог определения фона снижается (255→240→220…), "
            "что позволяет программе захватить всё более серые поля."
        ))
        self.chk_rt = QCheckBox("Преобразовать прозрачный в белый")
        self.chk_pdf = QCheckBox("PDF в квадрат")
        self.chk_pdf.setChecked(True)
        bpdf = QToolButton()
        bpdf.setText("ℹ")
        bpdf.setStyleSheet("border:none;color:blue;")
        bpdf.clicked.connect(lambda: QMessageBox.information(
            self, "Справка",
            "1-стр PDF → фото в квадрате. Много-стр PDF → просто скачается и скопируется в обработанные как есть."
        ))
        
        sub_layout.addWidget(self.chk_cs)
        sub_layout.addSpacing(4)
        sub_layout.addWidget(QLabel("Отступ %:"))
        sub_layout.addWidget(self.pi)
        sub_layout.addWidget(bpi)
        sub_layout.addSpacing(12)
        sub_layout.addWidget(self.chk_rt)
        sub_layout.addSpacing(12)
        sub_layout.addWidget(self.chk_pdf)
        sub_layout.addWidget(bpdf)
        sub_layout.addStretch()
        
        wl.addWidget(self.white_bg_container)
        ml.addWidget(gw)

    def _on_white_bg_toggled(self):
        enabled = self.chk_process_white.isChecked()
        self.white_bg_container.setEnabled(enabled)
        # Визуально "тушим" контейнер если выключено
        if enabled:
            self.white_bg_container.setStyleSheet("")
        else:
            self.white_bg_container.setStyleSheet("opacity: 0.6;")

    def _build_log_section(self, ml):
        lh = QHBoxLayout()
        lh.addWidget(QLabel("🖥 Лог:"))
        lh.addStretch()
        self.bel = QToolButton()
        self.bel.setText("⛶")
        self.bel.clicked.connect(self.show_log_d)
        lh.addWidget(self.bel)
        ml.addLayout(lh)

        self.lt = QTableWidget()
        self.lt.setColumnCount(3)
        self.lt.setHorizontalHeaderLabels(["Статус", "Источник", "Сообщение"])
        self.lt.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.lt.horizontalHeader().setSectionResizeMode(1, QHeaderView.Interactive)
        self.lt.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
        self.lt.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.lt.setColumnWidth(1, 380)
        self.lt.setMinimumHeight(110)
        self.lt.setMaximumHeight(180)
        ml.addWidget(self.lt)

        self.ls2 = QLabel("✅ ОК: 0 | 📏 Размер: 0 | ❓ Белый: 0 | ❌ Ошибки: 0 | 💾 0.0 МБ | ⏱ 00:00")
        self.ls2.setStyleSheet("font-weight:bold;font-size:11px;")
        ml.addWidget(self.ls2)
        ml.addStretch(1)

    def _build_progress_section(self, ml):
        self.pw = ValeraProgressWidget()
        self.pw.setVisible(False)
        ml.addWidget(self.pw)

    def _build_buttons_section(self, ml):
        bl = QHBoxLayout()
        bl.setSpacing(6)
        self.btn_start = QPushButton("▶ ЗАПУСТИТЬ")
        self.btn_start.setMinimumHeight(50)
        self.btn_start.setStyleSheet(
            "QPushButton{font-weight:bold;font-size:16px;background-color:#4CA3E0;color:white;border-radius:5px}"
            "QPushButton:hover{background-color:#3C93D0}"
        )
        self.btn_rej = QPushButton("📦 Просмотр результатов")
        self.btn_rej.setMinimumHeight(50)
        self.btn_rej.setVisible(False)
        self.btn_rej.setStyleSheet(
            "QPushButton{font-weight:bold;font-size:14px;background-color:#27AE60;color:white;border-radius:5px}"
            "QPushButton:hover{background-color:#229954}"
        )
        self.btn_cancel = QPushButton("✖ ОТМЕНА")
        self.btn_cancel.setMinimumHeight(50)
        self.btn_cancel.setVisible(False)
        self.btn_cancel.setStyleSheet(
            "QPushButton{font-weight:bold;font-size:16px;background-color:#E74C3C;color:white;border-radius:5px}"
        )
        bl.addWidget(self.btn_start, 3)
        bl.addWidget(self.btn_rej, 2)
        bl.addWidget(self.btn_cancel, 1)
        ml.addLayout(bl)

        self.bs.clicked.connect(self.pick_s)
        self.bd.clicked.connect(self.pick_d)
        self.btn_start.clicked.connect(self.toggle)
        self.btn_cancel.clicked.connect(self.cancel)
        self.btn_rej.clicked.connect(self.open_rejected)
        
        # Подключаем сигналы для отслеживания изменений
        for w in [self.fb, self.al, self.mi, self.ma, self.up, self.chk_cs, self.pi, 
                  self.chk_rt, self.chk_pdf, self.chk_process_white, self.chk_rm]:
            if isinstance(w, QComboBox):
                w.activated.connect(self.on_pm)
            elif isinstance(w, QCheckBox):
                w.clicked.connect(self.on_pm)
            else:
                w.textEdited.connect(self.on_pm)

    def _build_author_section(self, ml):
        bottom_row = QHBoxLayout()
        bottom_row.addStretch()
        self.lbl_author = QLabel('<a href="#" style="color: #999; text-decoration: none; font-size: 10px;">Автор</a>')
        self.lbl_author.setAlignment(Qt.AlignRight | Qt.AlignBottom)
        self.lbl_author.linkActivated.connect(self.show_author)
        bottom_row.addWidget(self.lbl_author)
        ml.addLayout(bottom_row)

    def show_author(self):
        AuthorDialog(self).exec()

    def pick_watermark(self):
        f, _ = QFileDialog.getOpenFileName(self, "Выбрать водный знак", "", "Изображения (*.png *.jpg *.jpeg *.webp)")
        if f:
            self.watermark_path = f
            self.lbl_wm.setText(os.path.basename(f))
            self.lbl_wm.setStyleSheet("color:black;font-size:10px;")
            self.btn_wm_clear.setVisible(True)

    def clear_watermark(self):
        self.watermark_path = ""
        self.lbl_wm.setText("")
        self.btn_wm_clear.setVisible(False)

    def on_pm(self):
        if self._ap:
            return
        if self.pc.currentText() != "Пользовательский":
            self.pc.blockSignals(True)
            self.pc.setCurrentText("Пользовательский")
            self.pc.blockSignals(False)

    def on_ps(self):
        self._ap = True
        self.apply_p(self.pc.currentText())
        self._ap = False

    def apply_p(self, n: str):
        if n not in PRESETS:
            return
        p = PRESETS[n]
        self.mi.setText(str(p.min_px))
        self.ma.setText(str(p.max_px))
        self.up.setText(str(p.max_upscale_pct))
        self.al.setCurrentText("По высоте" if p.align == "height" else "По ширине")
        self.fb.setCurrentText(p.fmt)
        self.chk_cs.setChecked(p.center_square)
        self.pi.setText(str(p.padding_pct))
        self.chk_rm.setChecked(p.remove_meta)
        self.chk_rt.setChecked(p.replace_transparent)
        self.chk_process_white.setChecked(p.process_white_bg)
        self._on_white_bg_toggled()

    def on_stc(self):
        st = self.stc.currentText()
        self.weo.setVisible(st == "Excel")
        self.wer.setVisible(st == "Excel")
        self.source_path = None
        self.ls.setText("Не выбран")
        self.ls.setStyleSheet("color:gray;")
        self.btn_start.setText("▶ ЗАПУСТИТЬ")

    def dragEnterEvent(self, e):
        if e.mimeData().hasUrls():
            e.acceptProposedAction()

    def dropEvent(self, e):
        p = e.mimeData().urls()[0].toLocalFile()
        if os.path.isfile(p) and p.lower().endswith((".xlsx", ".xls")):
            self.stc.setCurrentText("Excel")
            self.on_stc()
            self.set_s(p)
        elif os.path.isdir(p):
            self.stc.setCurrentText("Папка")
            self.on_stc()
            self.set_s(p)

    def set_s(self, p: str):
        self.source_path = p
        self.ls.setText(os.path.basename(p) if os.path.isfile(p) else p)
        self.ls.setStyleSheet("color:black;")

    def add_lr(self, s: str, so: str, m: str):
        self.log_buffer.append((s, so, m))

    def flush_log(self):
        if not self.log_buffer:
            return
        b = self.log_buffer[:]
        self.log_buffer.clear()
        self.lt.setUpdatesEnabled(False)
        for s, so, m in b:
            r = self.lt.rowCount()
            self.lt.insertRow(r)
            for c, t in enumerate([s, so, m]):
                self.lt.setItem(r, c, QTableWidgetItem(t))
        self.lt.setUpdatesEnabled(True)
        if not (self.log_dialog and self.log_dialog.isVisible()):
            self.lt.scrollToBottom()
        if self.log_dialog and self.log_dialog.isVisible():
            self.log_dialog.add_rows_batch(b)

    def show_log_d(self):
        if not self.log_dialog:
            self.log_dialog = LogDialog(self)
            rows = []
            for r in range(self.lt.rowCount()):
                rows.append((
                    self.lt.item(r, 0).text(),
                    self.lt.item(r, 1).text(),
                    self.lt.item(r, 2).text()
                ))
            self.log_dialog.add_rows_batch(rows)
        self.log_dialog.show()
        self.log_dialog.raise_()

    def show_info(self):
        QMessageBox.information(
            self, "Инструкция",
            "Общий алгоритм:\n"
            "1. Выберите источник (Excel с ссылками или Папка с фото).\n"
            "2. Укажите место назначения.\n"
            "3. Настройте параметры или выберите пресет.\n"
            "4. Нажмите ЗАПУСТИТЬ.\n"
            "5. После завершения нажмите 'Просмотр результатов' для ручной доработки фото.\n"
            "6. 'Обрезать в квадрат' — срезает лишний фон, но фото останется квадратным. "
            "Обрезанные края объекта (прямые линии) будут прижаты к краю, а свободное место заполнится белым.\n"
            "7. Многостраничные PDF копируются как есть, одностраничные — конвертируются в фото."
        )

    def pick_s(self):
        if self.stc.currentText() == "Excel":
            f, _ = QFileDialog.getOpenFileName(self, "Excel", "", "Excel (*.xlsx *.xls)")
            if f and os.path.exists(f):
                self.set_s(f)
        else:
            f = QFileDialog.getExistingDirectory(self, "Папка с фото")
            if f:
                self.set_s(f)

    def pick_d(self):
        f = QFileDialog.getExistingDirectory(self, "Назначение")
        if f:
            self.output_dir = f
            self.ld.setText(f)
            self.ld.setStyleSheet("color:black;")

    def gather(self) -> AppSettings:
        try:
            return AppSettings(
                source=self.source_path or "",
                source_type=self.stc.currentText(),
                out_dir=self.output_dir or os.path.join(os.path.expanduser("~"), "Desktop"),
                article_col=self.ca.currentText().upper(),
                url_from=self.cuf.currentText().upper(),
                url_to=self.cut.currentText().upper(),
                rename_mode="article" if self.cr.currentText() == "По артикулу" else "original",
                folder_sort=self.cfs.currentText(),
                min_px=int(self.mi.text()),
                max_px=int(self.ma.text()),
                max_upscale_pct=int(self.up.text()),
                align="height" if self.al.currentText() == "По высоте" else "width",
                fmt=self.fb.currentText(),
                center_square=self.chk_cs.isChecked(),
                padding_pct=int(self.pi.text()),
                replace_transparent=self.chk_rt.isChecked(),
                remove_meta=self.chk_rm.isChecked(),
                report_copy=self.chk_rc.isChecked(),
                clean_raw=self.chk_cr.isChecked(),
                ssl_verify=self.chk_ssl.isChecked(),
                aggressive_parse=self.chk_agg.isChecked(),
                preset_name=self.pc.currentText(),
                pdf_always_square=self.chk_pdf.isChecked(),
                selected_rejected_files=self._srf,
                white_threshold=max(80, Defaults.THRESHOLD_WHITE - (self.force_threshold_level * 20)),
                watermark_path=self.watermark_path,
                process_white_bg=self.chk_process_white.isChecked()
            )
        except ValueError:
            raise ValueError("Ошибка в числах!")

    def toggle(self):
        if self.worker and self.worker.isRunning():
            if self.worker._p:
                self.worker.resume()
                self.btn_start.setText("⏸ ОСТАНОВИТЬ")
            else:
                self.worker.pause()
                self.btn_start.setText("▶ ПРОДОЛЖИТЬ")
            return
        if not self.source_path:
            QMessageBox.warning(self, "Ошибка", "Выберите источник")
            return
        if not self.output_dir:
            if QMessageBox.question(self, "Внимание", "Сохранить на Рабочий стол?") == QMessageBox.No:
                return
            self.output_dir = os.path.join(os.path.expanduser("~"), "Desktop")
        if self.stc.currentText() == "Excel" and is_excel_locked(self.source_path):
            QMessageBox.warning(self, "Внимание", "Закройте Excel!")
            return
        try:
            s = self.gather()
        except ValueError as e:
            QMessageBox.warning(self, "Ошибка", str(e))
            return
        self.force_threshold_level = 0
        self.start_w(s)

    def start_w(self, s: AppSettings):
        self.lt.setRowCount(0)
        self.pw.setVisible(True)
        self.pw.setValue(0)
        self.btn_cancel.setVisible(True)
        self.btn_rej.setVisible(False)
        self.btn_start.setText("⏸ ОСТАНОВИТЬ")
        self.worker = Worker(s)
        self.worker.progress.connect(self.upd_p)
        self.worker.stats_updated.connect(self.upd_s)
        self.worker.log_row.connect(self.add_lr)
        self.worker.finished.connect(self.done)
        self.worker.error.connect(self.fail)
        self.worker.download_errors_ready.connect(self.on_download_errors)
        self.worker.start()

    def cancel(self):
        if self.worker and self.worker.isRunning():
            self.worker.cancel()
            self.btn_cancel.setEnabled(False)

    def upd_p(self, v: int):
        self.pw.setValue(v)

    def upd_s(self, ok, ds, dw, f, b, t):
        mins = int(t) // 60
        secs = int(t) % 60
        self.ls2.setText(
            f"✅ ОК: {ok} | 📏 Размер: {ds} | ❓ Белый: {dw} | ❌ Ошибки: {f} | "
            f"💾 {b/(1024*1024):.1f} МБ | ⏱ {mins:02d}:{secs:02d}"
        )

    def on_download_errors(self, errors):
        self.last_download_errors = errors

    def done(self, stats: dict):
        self.pw.setValue(100)
        self.btn_start.setText("▶ ЗАПУСТИТЬ")
        self.btn_cancel.setVisible(False)
        self.btn_cancel.setEnabled(True)
        self.save_s()
        self.last_processed_dir = stats.get("processed_dir")
        self.add_lr(
            "[СИСТЕМА]", "Готово",
            f"Время {int(stats.get('time', 0))//60:02d}:{int(stats.get('time', 0))%60:02d}"
        )
        self.flush_log()
        has_issues = (
            stats.get("defect_white", 0) > 0 or
            stats.get("defect_size", 0) > 0 or
            bool(self.last_download_errors)
        )
        self.btn_rej.setVisible(True)
        if has_issues:
            self.btn_rej.setText("👀 Просмотр")
            self.btn_rej.setStyleSheet(
                "QPushButton{font-weight:bold;font-size:14px;background-color:#FF9800;color:white;border-radius:5px}"
                "QPushButton:hover{background-color:#F57C00}"
            )
        else:
            self.btn_rej.setText("👀 Просмотр")
            self.btn_rej.setStyleSheet(
                "QPushButton{font-weight:bold;font-size:14px;background-color:#27AE60;color:white;border-radius:5px}"
                "QPushButton:hover{background-color:#229954}"
            )

    def fail(self, msg: str):
        self.pw.setVisible(False)
        self.btn_start.setText("▶ ЗАПУСТИТЬ")
        self.btn_cancel.setVisible(False)
        self.add_lr("[ОШИБКА]", "Критическая", msg[:300])
        self.flush_log()
        QMessageBox.critical(self, "Ошибка", msg)

    def open_rejected(self):
        target_dir = self.last_processed_dir
        if not target_dir or not os.path.exists(target_dir):
            if not self.output_dir or not os.path.exists(self.output_dir):
                return
            dirs = [
                os.path.join(self.output_dir, d)
                for d in os.listdir(self.output_dir)
                if d.startswith("обработано_")
            ]
            if dirs:
                target_dir = max(dirs, key=os.path.getmtime)
            else:
                goto = os.path.join(self.output_dir, "Готово")
                if os.path.exists(goto):
                    target_dir = goto
                else:
                    return
        try:
            s = self.gather()
        except Exception:
            return
        dlg = ProductPreviewDialog(
            target_dir, s, self,
            download_errors=self.last_download_errors,
            threshold_level=self.force_threshold_level
        )
        dlg.exec()

    def on_rej_sel(self, files):
        self._srf = files
        if not files:
            return
        try:
            self.force_threshold_level += 1
            s = self.gather()
            s.source_type = "Отбракованное"
        except Exception:
            return
        self.start_w(s)

    def save_s(self):
        s = self.settings_io
        s.setValue("source", self.source_path or "")
        s.setValue("out_dir", self.output_dir or "")
        s.setValue("st", self.stc.currentText())
        s.setValue("art", self.ca.currentText())
        s.setValue("uf", self.cuf.currentText())
        s.setValue("ut", self.cut.currentText())
        s.setValue("rm", self.cr.currentText())
        s.setValue("fs", self.cfs.currentText())
        s.setValue("mi", self.mi.text())
        s.setValue("ma", self.ma.text())
        s.setValue("up", self.up.text())
        s.setValue("fmt", self.fb.currentText())
        s.setValue("al", self.al.currentText())
        s.setValue("meta", self.chk_rm.isChecked())
        s.setValue("cs", self.chk_cs.isChecked())
        s.setValue("pp", self.pi.text())
        s.setValue("rt", self.chk_rt.isChecked())
        s.setValue("rc", self.chk_rc.isChecked())
        s.setValue("clr", self.chk_cr.isChecked())
        s.setValue("ssl", self.chk_ssl.isChecked())
        s.setValue("agg", self.chk_agg.isChecked())
        s.setValue("pr", self.pc.currentText())
        s.setValue("pdf", self.chk_pdf.isChecked())
        s.setValue("pwb", self.chk_process_white.isChecked())

    def load_settings(self):
        s = self.settings_io
        self.source_path = s.value("source", "")
        self.output_dir = s.value("out_dir", "")
        self.stc.setCurrentText(s.value("st", "Excel"))
        self.on_stc()
        if self.source_path and os.path.exists(self.source_path):
            self.ls.setText(os.path.basename(self.source_path))
            self.ls.setStyleSheet("color:black;")
        if self.output_dir and os.path.exists(self.output_dir):
            self.ld.setText(self.output_dir)
            self.ld.setStyleSheet("color:black;")
        self.ca.setCurrentText(s.value("art", "A"))
        self.cuf.setCurrentText(s.value("uf", "B"))
        self.cut.setCurrentText(s.value("ut", "P"))
        self.cr.setCurrentText(s.value("rm", "По артикулу"))
        self.cfs.setCurrentText(s.value("fs", "В порядке Excel"))
        self.mi.setText(s.value("mi", "0"))
        self.ma.setText(s.value("ma", "4000"))
        self.up.setText(s.value("up", "50"))
        self.fb.setCurrentText(s.value("fmt", "jpg"))
        self.al.setCurrentText(s.value("al", "По высоте"))
        self.chk_rm.setChecked(s.value("meta", False, type=bool))
        self.chk_cs.setChecked(s.value("cs", True, type=bool))
        self.pi.setText(s.value("pp", "10"))
        self.chk_rt.setChecked(s.value("rt", True, type=bool))
        self.chk_rc.setChecked(s.value("rc", False, type=bool))
        self.chk_cr.setChecked(s.value("clr", False, type=bool))
        self.chk_ssl.setChecked(s.value("ssl", True, type=bool))
        self.chk_agg.setChecked(s.value("agg", True, type=bool))
        self.chk_pdf.setChecked(s.value("pdf", True, type=bool))
        self.chk_process_white.setChecked(s.value("pwb", True, type=bool))
        self._on_white_bg_toggled()
        pr = s.value("pr", "santehnica.ru")
        if pr in PRESETS:
            self._ap = True
            self.pc.setCurrentText(pr)
            self.apply_p(pr)
            self._ap = False


# ============================================================
# ENTRY POINT
# ============================================================
def global_exception_handler(t, v, b):
    try:
        err_msg = "".join(traceback.format_exception(t, v, b))
        LOGGER.critical(err_msg)
        QMessageBox.critical(
            None, "Критическая ошибка",
            f"Произошла ошибка:\n\n{err_msg[:500]}\n\nПодробности в логе."
        )
    except Exception as e:
        print(f"Exception in global exception handler: {e}")


if sys.platform == "win32":
    import ctypes
    ctypes.windll.kernel32.FreeConsole()
    
if __name__ == "__main__":
    sys.excepthook = global_exception_handler
    app = QApplication(sys.argv)
    setup_logging()
    apply_light_theme(app)
    w = MainWindow()
    w.show()
    sys.exit(app.exec())
