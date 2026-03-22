import sys, os, io, json, requests, warnings, traceback
from urllib.parse import urlparse, unquote
from PySide6.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QFileDialog, QLineEdit, QCheckBox, QProgressBar, QComboBox, QRadioButton, 
    QMessageBox, QGridLayout, QSizePolicy, QSpacerItem, QToolButton, QStyle, QFileIconProvider
)
from PySide6.QtCore import QThread, Signal, QSettings, Qt, QRect, QSize, QFileInfo
from PySide6.QtGui import QPixmap, QPainter, QColor, QFont, QBrush, QPen, QIcon
from PIL import Image as PILImage
import openpyxl
from openpyxl.styles import PatternFill

# Отключаем предупреждения SSL
warnings.filterwarnings("ignore", message="Unverified HTTPS request")

# ---------- Виджет с Валерой (Прогресс) ----------
class ValeraProgressWidget(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        # Увеличили высоту, чтобы совпадала с кнопкой Запустить
        self.setMinimumHeight(50) 
        self._value = 0
        self.valera_pixmap = None
        
        # Безопасная загрузка картинки
        try:
            script_dir = os.path.dirname(os.path.abspath(__file__))
            valera_path = os.path.join(script_dir, "Valera.png")
            if os.path.exists(valera_path):
                self.valera_pixmap = QPixmap(valera_path)
        except:
            pass
            
        self.bg_color = QColor(230, 230, 230)
        self.fill_color = QColor(76, 163, 224)

    def setValue(self, value):
        self._value = value
        self.update()

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)
        
        # 1. Фон (серый)
        painter.setBrush(QBrush(self.bg_color))
        painter.setPen(Qt.NoPen)
        painter.drawRoundedRect(0, 0, self.width(), self.height(), 5, 5)
        
        # 2. Заполнение (синий)
        if self._value > 0:
            width = int(self.width() * (min(self._value, 100) / 100.0))
            painter.setBrush(QBrush(self.fill_color))
            painter.drawRoundedRect(0, 0, width, self.height(), 5, 5)
            
            # 3. Валера (Толстый и красивый)
            if self.valera_pixmap and not self.valera_pixmap.isNull():
                p_height = self.height() - 4
                scaled_pix = self.valera_pixmap.scaledToHeight(p_height, Qt.SmoothTransformation)
                p_width = scaled_pix.width()
                
                if self._value >= 100:
                    x_pos = self.width() - p_width - 10
                else:
                    x_pos = width - p_width // 2
                    if x_pos < 0: x_pos = 0
                
                painter.drawPixmap(x_pos, 2, scaled_pix)

        # 4. Текст
        painter.setPen(QPen(QColor(255, 255, 255)))
        font = painter.font()
        font.setBold(True)
        font.setPointSize(12)
        painter.setFont(font)
        text = f"Ход работ: {self._value}%"
        painter.drawText(QRect(0, 0, self.width(), self.height()), Qt.AlignCenter, text)

# ---------- Utils ----------
def parse_columns(text):
    text = text.replace(" ", "").upper()
    cols = []
    for part in text.split(","):
        if "-" in part:
            start, end = part.split("-")
            start_idx = ord(start)-65
            end_idx = ord(end)-65
            cols.extend(range(start_idx, end_idx+1))
        else:
            cols.append(ord(part)-65)
    return cols

# ---------- Worker ----------
class Worker(QThread):
    progress = Signal(int)
    finished = Signal(int, int)
    error = Signal(str)

    def __init__(self, settings):
        super().__init__()
        self.s = settings

    def run(self):
        try:
            ok_rows, fail_rows = process_files(self.s, self.progress)
            self.finished.emit(ok_rows, fail_rows)
        except Exception as e:
            self.error.emit(str(e))

# ---------- Main App ----------
class App(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Фотогрузчик Валера")
        self.setFixedSize(600, 420) # Уменьшили размеры окна (было 650x550)
        
        # Иконка окна
        try:
            script_dir = os.path.dirname(os.path.abspath(__file__))
            valera_path = os.path.join(script_dir, "Valera.png")
            if os.path.exists(valera_path):
                self.setWindowIcon(QIcon(valera_path))
        except: pass
        
        self.excel_path = None
        self.output_dir = None
        self.wm_path = None
        self.settings_io = QSettings("ValeraSoft", "PhotoLoader")
        
        # Провайдер иконок
        self.icon_provider = QFileIconProvider()

        # Главный вертикальный слой
        main_layout = QVBoxLayout()
        main_layout.setContentsMargins(10, 10, 10, 10) # Уменьшили отступы (было 15)
        
        # 1. Верхняя часть (Настройки)
        grid = QGridLayout()
        grid.setSpacing(5) 

        # --- Row 0: Excel ---
        self.check_excel = QLabel("❌")
        self.check_excel.setFixedWidth(20)
        self.btn_excel = QPushButton(" Выбрать Excel")
        self.btn_excel.setFixedWidth(130) # Уменьшили ширину кнопки (было ~200)
        self.btn_excel.setIcon(self.style().standardIcon(QStyle.SP_FileIcon))
        
        self.btn_open_excel = QToolButton()
        self.btn_open_excel.setIcon(self.style().standardIcon(QStyle.SP_FileIcon))
        self.btn_open_excel.setToolTip("Открыть Excel")
        self.btn_open_excel.setFixedSize(30, 30)
        
        self.label_excel = QLabel("Файл не выбран")
        self.label_excel.setStyleSheet("color: gray;")
        
        grid.addWidget(self.check_excel, 0, 0)
        grid.addWidget(self.btn_excel, 0, 1)
        grid.addWidget(self.label_excel, 0, 2)
        grid.addWidget(self.btn_open_excel, 0, 3)

        # --- Row 1: Folder ---
        self.check_dir = QLabel("❌")
        self.btn_dir = QPushButton(" Выбрать папку")
        self.btn_dir.setFixedWidth(130) # Уменьшили ширину
        self.btn_dir.setIcon(self.style().standardIcon(QStyle.SP_DirIcon))
        self.label_dir = QLabel("Папка не выбрана")
        self.label_dir.setStyleSheet("color: gray;")
        
        grid.addWidget(self.check_dir, 1, 0)
        grid.addWidget(self.btn_dir, 1, 1)
        grid.addWidget(self.label_dir, 1, 2, 1, 2)

        # --- Row 2: Watermark ---
        self.check_wm = QLabel("❌")
        self.btn_wm = QPushButton(" Водяной знак")
        self.btn_wm.setFixedWidth(130) # Уменьшили ширину
        self.btn_wm.setIcon(self.style().standardIcon(QStyle.SP_FileIcon))
        self.label_wm = QLabel("Не выбран")
        self.label_wm.setStyleSheet("color: gray;")
        
        grid.addWidget(self.check_wm, 2, 0)
        grid.addWidget(self.btn_wm, 2, 1)
        grid.addWidget(self.label_wm, 2, 2, 1, 2)

        # --- Row 3: Columns ---
        h_cols = QHBoxLayout()
        self.article_input = QLineEdit("A")
        self.article_input.setFixedWidth(40)
        self.urls_input = QLineEdit("B-P")
        self.urls_input.setFixedWidth(80)
        
        h_cols.addWidget(QLabel("Артикул:"))
        h_cols.addWidget(self.article_input)
        h_cols.addSpacing(15)
        h_cols.addWidget(QLabel("Колонки ссылок:"))
        h_cols.addWidget(self.urls_input)
        h_cols.addStretch()
        grid.addLayout(h_cols, 3, 0, 1, 4)

        # --- Row 4: Size & Align ---
        h_size = QHBoxLayout()
        self.radio_height = QRadioButton("По высоте")
        self.radio_width = QRadioButton("По ширине")
        self.radio_height.setChecked(True)
        self.size_input = QLineEdit("800")
        self.size_input.setFixedWidth(50)
        
        h_size.addWidget(QLabel("Выравнивание:"))
        h_size.addWidget(self.radio_height)
        h_size.addWidget(self.radio_width)
        h_size.addSpacing(15)
        h_size.addWidget(QLabel("Размер px:"))
        h_size.addWidget(self.size_input)
        h_size.addStretch()
        grid.addLayout(h_size, 4, 0, 1, 4)

        # --- Row 5: Format & Crop & Rename ---
        h_opts = QHBoxLayout()
        self.check_crop = QCheckBox("Квадрат")
        self.format_box = QComboBox()
        self.format_box.addItems(["jpg","png","webp"])
        self.format_box.setFixedWidth(65)
        
        # Новая галочка переименования
        self.check_rename = QCheckBox("Переименовывать по артикулу")
        self.check_rename.setChecked(True)
        self.check_rename.setToolTip("Если включено: Артикул_1.jpg\nЕсли выключено: сохраняется имя файла из ссылки.")
        
        h_opts.addWidget(self.check_crop)
        h_opts.addSpacing(10)
        h_opts.addWidget(QLabel("Формат:"))
        h_opts.addWidget(self.format_box)
        h_opts.addSpacing(10)
        h_opts.addWidget(self.check_rename)
        h_opts.addStretch()
        grid.addLayout(h_opts, 5, 0, 1, 4)

        # --- Row 6: Rejection & Min Size ---
        h_reject = QHBoxLayout()
        
        self.check_reject = QCheckBox("Отбраковывать при ошибке")
        self.check_reject.setChecked(True)
        
        self.btn_info = QToolButton()
        self.btn_info.setText("ℹ")
        self.btn_info.setStyleSheet("border: none; color: blue; font-weight: bold; font-size: 14px;")
        self.btn_info.clicked.connect(self.show_reject_info)
        
        self.min_size_combo = QComboBox()
        self.min_size_combo.addItems(["Без ограничений","5 KB","50 KB","100 KB"])
        self.min_size_combo.setFixedWidth(120)
        
        h_reject.addWidget(self.check_reject)
        h_reject.addWidget(self.btn_info)
        h_reject.addSpacing(15)
        h_reject.addWidget(QLabel("Мин. размер файла:"))
        h_reject.addWidget(self.min_size_combo)
        h_reject.addStretch()
        grid.addLayout(h_reject, 6, 0, 1, 4)

        main_layout.addLayout(grid)

        # --- Растягиватель ---
        main_layout.addStretch(1)

        # --- Progress Bar (Над кнопкой) ---
        self.progress_widget = ValeraProgressWidget()
        self.progress_widget.setVisible(False) 
        main_layout.addWidget(self.progress_widget)

        # --- Start Button (Fixed at bottom) ---
        self.btn_start = QPushButton("ЗАПУСТИТЬ")
        self.btn_start.setMinimumHeight(50) # Такая же высота как у прогресс-бара
        self.btn_start.setStyleSheet("""
            QPushButton {
                font-weight: bold; 
                font-size: 16px; 
                background-color: #4CA3E0; 
                color: white; 
                border-radius: 5px;
            }
            QPushButton:hover { background-color: #3C93D0; }
            QPushButton:disabled { background-color: #cccccc; }
        """)
        main_layout.addWidget(self.btn_start)

        self.status = QLabel("")
        main_layout.addWidget(self.status)

        # --- Footer ---
        h_footer = QHBoxLayout()
        h_footer.addStretch()
        self.btn_author = QPushButton(" Автор")
        self.btn_author.setFlat(True)
        self.btn_author.setStyleSheet("color: gray; border: none;")
        self.btn_author.clicked.connect(self.show_author)
        h_footer.addWidget(self.btn_author)
        main_layout.addLayout(h_footer)

        self.setLayout(main_layout)
        self.load_settings_ui()

        # Connections
        self.btn_excel.clicked.connect(self.pick_excel)
        self.btn_open_excel.clicked.connect(self.open_excel)
        self.btn_dir.clicked.connect(self.pick_dir)
        self.btn_wm.clicked.connect(self.pick_wm)
        self.btn_start.clicked.connect(self.start)

    def show_reject_info(self):
        QMessageBox.information(self, "Параметры отбраковки",
            "Если галочка стоит, система проверяет каждый файл по критериям:\n\n"
            "1. Ошибка скачивания: файл не удалось скачать.\n"
            "2. Размер файла: если файл меньше указанного 'Мин. размера файла'.\n\n"
            "Если файл не проходит проверку, строка в Excel краснеет, а файл (если скачался) "
            "попадает в папку FAIL вместо OK."
        )

    def check_and_fix_merged_cells(self, filepath):
        try:
            wb = openpyxl.load_workbook(filepath)
            has_merges = False
            
            for sheet in wb.worksheets:
                if sheet.merged_cells.ranges:
                    has_merges = True
                    break
            
            if has_merges:
                msgBox = QMessageBox()
                msgBox.setIcon(QMessageBox.Warning)
                msgBox.setWindowTitle("Внимание: Объединенные ячейки")
                msgBox.setText("В файле обнаружены объединенные ячейки. Это может привести к ошибкам.\n\n"
                    "Хотите автоматически снять объединение?")
                
                btn_yes = msgBox.addButton("Да", QMessageBox.YesRole)
                btn_no = msgBox.addButton("Нет", QMessageBox.NoRole)
                
                msgBox.exec()
                
                if msgBox.clickedButton() == btn_yes:
                    for sheet in wb.worksheets:
                        merged_ranges = list(sheet.merged_cells.ranges)
                        for merge_range in merged_ranges:
                            min_col = merge_range.min_col
                            min_row = merge_range.min_row
                            max_col = merge_range.max_col
                            max_row = merge_range.max_row
                            
                            top_left_value = sheet.cell(row=min_row, column=min_col).value
                            
                            sheet.unmerge_cells(str(merge_range))
                            
                            for r in range(min_row, max_row + 1):
                                for c in range(min_col, max_col + 1):
                                    sheet.cell(row=r, column=c, value=top_left_value)
                    
                    wb.save(filepath)
                    QMessageBox.information(self, "Успех", "Объединение ячеек снято.")
                    return True
            return False
        except Exception as e:
             QMessageBox.warning(self, "Ошибка", f"Не удалось проверить объединенные ячейки:\n{e}")
             return False

    def load_settings_ui(self):
        self.excel_path = self.settings_io.value("excel", "")
        self.output_dir = self.settings_io.value("out_dir", "")
        self.wm_path = self.settings_io.value("wm", "")
        
        if self.excel_path and os.path.exists(self.excel_path):
            self.check_excel.setText("✅")
            self.label_excel.setText(os.path.basename(self.excel_path))
            self.label_excel.setStyleSheet("color: black;")
            icon = self.icon_provider.icon(QFileInfo(self.excel_path))
            self.btn_excel.setIcon(icon)
            self.btn_open_excel.setIcon(icon)
            
        if self.output_dir and os.path.exists(self.output_dir):
            self.check_dir.setText("✅")
            self.label_dir.setText(self.output_dir)
            self.label_dir.setStyleSheet("color: black;")
            icon = self.icon_provider.icon(QFileInfo(self.output_dir))
            self.btn_dir.setIcon(icon)
            
        if self.wm_path and os.path.exists(self.wm_path):
            self.check_wm.setText("✅")
            self.label_wm.setText(os.path.basename(self.wm_path))
            self.label_wm.setStyleSheet("color: black;")
            icon = self.icon_provider.icon(QFileInfo(self.wm_path))
            self.btn_wm.setIcon(icon)
            
        self.article_input.setText(self.settings_io.value("article_col", "A"))
        self.urls_input.setText(self.settings_io.value("url_cols", "B-P"))
        self.size_input.setText(self.settings_io.value("size", "800"))
        self.check_crop.setChecked(self.settings_io.value("crop", False, type=bool))
        self.check_reject.setChecked(self.settings_io.value("reject", True, type=bool))
        self.check_rename.setChecked(self.settings_io.value("rename", True, type=bool))
        
        align = self.settings_io.value("align", "height")
        if align == "height": self.radio_height.setChecked(True)
        else: self.radio_width.setChecked(True)
        
        min_sz = self.settings_io.value("min_size", "Без ограничений")
        idx = self.min_size_combo.findText(min_sz)
        if idx >= 0: self.min_size_combo.setCurrentIndex(idx)
            
        fmt = self.settings_io.value("format", "jpg")
        idx_f = self.format_box.findText(fmt)
        if idx_f >= 0: self.format_box.setCurrentIndex(idx_f)

    def save_settings_ui(self):
        self.settings_io.setValue("excel", self.excel_path or "")
        self.settings_io.setValue("out_dir", self.output_dir or "")
        self.settings_io.setValue("wm", self.wm_path or "")
        self.settings_io.setValue("article_col", self.article_input.text())
        self.settings_io.setValue("url_cols", self.urls_input.text())
        self.settings_io.setValue("size", self.size_input.text())
        self.settings_io.setValue("crop", self.check_crop.isChecked())
        self.settings_io.setValue("reject", self.check_reject.isChecked())
        self.settings_io.setValue("rename", self.check_rename.isChecked())
        self.settings_io.setValue("align", "height" if self.radio_height.isChecked() else "width")
        self.settings_io.setValue("min_size", self.min_size_combo.currentText())
        self.settings_io.setValue("format", self.format_box.currentText())

    # ---------- Actions ----------
    def pick_excel(self):
        file,_ = QFileDialog.getOpenFileName(self, "Выберите Excel", "", "Excel Files (*.xlsx *.xls);;All Files (*)")
        if file and os.path.exists(file):
            self.excel_path = file
            self.label_excel.setText(os.path.basename(file))
            self.check_excel.setText("✅")
            self.label_excel.setStyleSheet("color: black;")
            icon = self.icon_provider.icon(QFileInfo(file))
            self.btn_excel.setIcon(icon)
            self.btn_open_excel.setIcon(icon)
            self.check_and_fix_merged_cells(file)

    def open_excel(self):
        if self.excel_path and os.path.exists(self.excel_path):
            try: os.startfile(self.excel_path)
            except Exception as e: QMessageBox.critical(self,"Ошибка",f"Не удалось открыть файл:\n{e}")

    def pick_dir(self):
        folder=QFileDialog.getExistingDirectory(self,"Выберите папку")
        if folder:
            self.output_dir=folder
            self.label_dir.setText(folder)
            self.check_dir.setText("✅")
            self.label_dir.setStyleSheet("color: black;")
            icon = self.icon_provider.icon(QFileInfo(folder))
            self.btn_dir.setIcon(icon)

    def pick_wm(self):
        file,_=QFileDialog.getOpenFileName(self,"Выберите PNG","*.png")
        if file:
            self.wm_path=file
            self.label_wm.setText(os.path.basename(file))
            self.check_wm.setText("✅")
            self.label_wm.setStyleSheet("color: black;")
            icon = self.icon_provider.icon(QFileInfo(file))
            self.btn_wm.setIcon(icon)

    def show_author(self):
        QMessageBox.information(self,"Автор",
            'Мой гитхаб https://github.com/allkirill/\n'
            'Пет-приют: vlookup-app.ru\n\n'
            'Фотоукладчик Валера всегда готов выйти на дело!', QMessageBox.Ok)

    # ---------- Start ----------
    def start(self):
        if not self.excel_path or not self.output_dir:
            QMessageBox.warning(self,"Ошибка","Выберите Excel и папку для сохранения")
            return
        
        try:
            sz = int(self.size_input.text())
        except:
            QMessageBox.warning(self,"Ошибка","Размер должен быть числом")
            return

        align="height" if self.radio_height.isChecked() else "width"
        min_size_map={"Без ограничений":0,"5 KB":5000,"50 KB":50000,"100 KB":100000}
        
        self.settings={
            "excel":self.excel_path,
            "out_dir":self.output_dir,
            "size":sz,
            "crop":self.check_crop.isChecked(),
            "wm":self.wm_path,
            "article_col":self.article_input.text(),
            "url_cols":self.urls_input.text(),
            "reject":self.check_reject.isChecked(),
            "rename": self.check_rename.isChecked(),
            "min_size":min_size_map[self.min_size_combo.currentText()],
            "format":self.format_box.currentText(),
            "align":align
        }
        
        self.btn_start.setEnabled(False)
        self.status.setText("")
        
        self.progress_widget.setVisible(True)
        self.progress_widget.setValue(0)
        
        self.worker=Worker(self.settings)
        self.worker.progress.connect(self.update_progress)
        self.worker.finished.connect(self.done)
        self.worker.error.connect(self.fail)
        self.worker.start()

    def update_progress(self, val):
        self.progress_widget.setValue(val)

    def done(self, ok_rows, fail_rows):
        self.progress_widget.setValue(100)
        self.btn_start.setEnabled(True)
        
        self.status.setText(f"Готово! Успешно: {ok_rows}, Отбраковано: {fail_rows}")
        self.save_settings_ui()
        QMessageBox.information(self,"Результат",
            f"Успешно: {ok_rows}\nОтбраковано: {fail_rows}\n\nСм. Excel (цветовые пометки)")

    def fail(self,msg):
        self.progress_widget.setVisible(False)
        self.btn_start.setEnabled(True)
        self.status.setText("Ошибка")
        QMessageBox.critical(self,"Ошибка",msg)

# ---------- Process Files ----------
def process_files(s, progress_signal):
    try:
        wb=openpyxl.load_workbook(s["excel"])
    except PermissionError:
        raise Exception("Закройте Excel файл!")

    session=requests.Session()
    session.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8",
        "Accept-Language": "ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7"
    })

    article_col=ord(s["article_col"].upper())-65
    url_cols=parse_columns(s["url_cols"])

    ok_dir=os.path.join(s["out_dir"],"OK")
    fail_dir=os.path.join(s["out_dir"],"FAIL")
    os.makedirs(ok_dir,exist_ok=True)
    os.makedirs(fail_dir,exist_ok=True)

    total_urls = 0
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for c in url_cols:
                if c < len(row) and row[c].value:
                    total_urls += 1
                    
    processed=0
    ok_rows=0
    fail_rows=0

    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            if article_col>=len(row): continue
            art=row[article_col].value
            if not art: continue
            
            row_error=False
            safe_art="".join(c if c.isalnum() else "_" for c in str(art))
            imgs=[]
            idx=1
            
            for col in url_cols:
                if col>=len(row) or not row[col].value: continue
                val=str(row[col].value).strip()
                
                try:
                    parsed_url = urlparse(val)
                    domain = f"{parsed_url.scheme}://{parsed_url.netloc}"
                    
                    resp = session.get(val, timeout=15, headers={"Referer": domain + "/"}, verify=False)
                    
                    if resp.status_code != 200:
                        raise Exception(f"Status {resp.status_code}")
                        
                    content = resp.content
                    
                    if s["min_size"] > 0 and len(content) < s["min_size"]:
                        raise Exception("File too small")

                    img=PILImage.open(io.BytesIO(content)).convert("RGBA")
                    
                    w,h=img.size
                    ratio=s["size"]/h if s["align"]=="height" else s["size"]/w
                    img=img.resize((int(w*ratio),int(h*ratio)))
                    
                    if s["crop"]:
                        m=min(img.size)
                        img=img.crop((0,0,m,m))
                    
                    # Логика имени файла
                    if s["rename"]:
                        filename = f"{safe_art}_{idx}.{s['format']}"
                    else:
                        url_path = unquote(parsed_url.path)
                        basename = os.path.basename(url_path)
                        name_part = os.path.splitext(basename)[0]
                        if not name_part: name_part = f"image_{idx}"
                        
                        safe_name = "".join(c if c.isalnum() or c in (' ', '_', '-') else "_" for c in name_part)
                        filename = f"{safe_name}.{s['format']}"
                        
                        # Проверка дубликатов
                        counter = 1
                        orig_filename = filename
                        while any(im[1] == filename for im in imgs):
                             name, ext = os.path.splitext(orig_filename)
                             filename = f"{name}_{counter}{ext}"
                             counter += 1

                    imgs.append((img, filename))
                    idx+=1
                    
                except Exception as e:
                    row_error=True
                
                processed+=1
                if total_urls>0: 
                    progress_signal.emit(int(processed/total_urls*100))
            
            target=fail_dir if (s["reject"] and row_error) else ok_dir
            
            if not row_error and imgs:
                fill_color="90EE90"
                ok_rows+=1
            elif row_error and imgs:
                fill_color="FFD700"
                fail_rows+=1 
            elif not imgs:
                fill_color="FF6347"
                fail_rows+=1
            else:
                fill_color="FFFFFF"
                
            for col in url_cols:
                if col<len(row):
                    row[col].fill=PatternFill(start_color=fill_color,fill_type="solid")
            
            folder=os.path.join(target,safe_art)
            os.makedirs(folder,exist_ok=True)
            
            for img_tuple in imgs:
                img = img_tuple[0]
                fname = img_tuple[1]
                
                path=os.path.join(folder, fname)
                
                if s["format"]=="jpg":
                    rgb=PILImage.new("RGB",img.size,(255,255,255))
                    rgb.paste(img,mask=img.split()[3])
                    rgb.save(path,quality=90)
                else:
                    img.save(path)
                    
    try:
        wb.save(s["excel"])
    except PermissionError:
         raise Exception("Не удалось сохранить Excel. Закройте файл!")

    return ok_rows, fail_rows

# ---------- Run with Error Handler ----------
def exception_hook(exctype, value, tb):
    error_msg = ''.join(traceback.format_exception(exctype, value, tb))
    print(error_msg)
    QMessageBox.critical(None, "Ошибка запуска", f"Программа упала с ошибкой:\n{error_msg}")

if __name__=="__main__":
    sys.excepthook = exception_hook
    app=QApplication(sys.argv)
    w=App()
    w.show()
    sys.exit(app.exec())